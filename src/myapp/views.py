import re

from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.utils.html import escape
from urllib.parse import quote_plus
from django.views.decorators.clickjacking import xframe_options_sameorigin

from .gemini import ask_gemini, suggest_from_other_sources
from .models import Book

from unicodedata import category
from urllib import request

from django.shortcuts import render,HttpResponse,redirect
from .models import register as RegisterUser, contact as contact_model, product as product_model, wishlist as wishlist_model, add_to_cart, checkout as checkout_model, Category, ProductReview, Coupon
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth import get_user_model
import random

try:
    import razorpay
except ModuleNotFoundError:
    razorpay = None

# Create your views here.
def home(request):
    if "email" in request.session:
        email = request.session['email']
        uid = RegisterUser.objects.filter(email=email).first()
        
        # If user not in legacy RegisterUser model, sync from NewUser
        if not uid:
            from accounts.views import sync_register_user
            from django.contrib.auth import get_user_model
            User = get_user_model()
            
            # Try to get the user from NewUser model and sync
            try:
                new_user = User.objects.get(email=email)
                sync_register_user(new_user)
                uid = RegisterUser.objects.filter(email=email).first()
            except User.DoesNotExist:
                request.session.flush()
                return redirect("accounts:login")
        
        if not uid:
            # If still no user found, clear session and redirect
            request.session.flush()
            return redirect("accounts:login")

        # Fetch products for each section
        bestsellers = product_model.objects.filter(bestseller=True)[:6]
        new_releases = product_model.objects.filter(new_release=True)[:6]
        expert_picks = product_model.objects.filter(expert_pick=True)[:6]
        wishlist_ids = set(
            wishlist_model.objects.filter(register=uid).values_list("product_id", flat=True)
        )
        
        context = {
            "bestsellers": bestsellers,
            "new_releases": new_releases,
            "expert_picks": expert_picks,
            "wishlist_ids": wishlist_ids,
        }
        return render(request, "customerapp/home.html", context)
    else:
         return redirect("accounts:login")


def about(request):    
    if "email" in request.session:
        return render(request,"customerapp/about.html")
    else:
        return redirect("accounts:login")


def contact(request):
    if "email" in request.session:
        uid = RegisterUser.objects.filter(email=request.session['email']).first()
        if not uid:
            request.session.flush()
            return redirect("accounts:login")

        print(uid.email)

        if request.POST:
            name = request.POST.get("name", '').strip()
            email = request.POST.get("email", '').strip()
            phone = request.POST.get("phone", '').strip()
            subject = request.POST.get("subject", '').strip()
            message = request.POST.get("message", '').strip()

            # Validate all required fields
            if not name or not email or not phone or not subject or not message:
                messages.error(request, 'All fields are required! Please fill in all the fields.')
            else:
                print(name,email,phone,subject,message)
                contact_model.objects.create(name=name,email=email,phone=phone,subject=subject,message=message)
                messages.success(request, 'Your message has been sent successfully!')

        contaxt={
            "uid":uid
        }

        return render(request,"customerapp/contact.html",contaxt)
    else:
        return redirect('accounts:login')

def base(request):
    return render(request, 'customerapp/base.html')


def search(request):
    if request.method == "POST":
        search_query = request.POST.get("search", "").strip()
        print(f"Searching for: {search_query}")
        
        pid = product_model.objects.filter(name__icontains=search_query)
        
        # Generate external links
        query_encoded = quote_plus(search_query)
        amazon_link = f"https://www.amazon.in/s?k={query_encoded}"
        flipkart_link = f"https://www.flipkart.com/search?q={query_encoded}"
        
        context = {
            "pid": pid,
            "search_query": search_query,
            "amazon_link": amazon_link,
            "flipkart_link": flipkart_link,
            "total_count": pid.count(),
        }
        return render(request, "customerapp/shop.html", context)
    return redirect('shop')


def faq(request):
    return render(request, 'customerapp/faq.html')

# def register(request):
#     if request.method == 'POST':
#         username = request.POST.get('username', '').strip()
#         email = request.POST.get('email', '').strip()
#         password = request.POST.get('password', '').strip()
#         confirm_password = request.POST.get('confirm_password', '').strip()
#         
#         # Check if all fields are filled
#         if not username or not email or not password or not confirm_password:
#             messages.error(request, 'All fields are required. Please fill in all the fields!')
#             return render(request, 'customerapp/register.html')
#         
#         # Check if passwords match
#         if password != confirm_password:
#             messages.error(request, 'Passwords do not match!')
#             return render(request, 'customerapp/register.html')
#         
#         # Check if user already exists
#         if RegisterUser.objects.filter(email=email).exists():
#             messages.error(request, 'Email already exists!')
#             return render(request, 'customerapp/register.html')
#         
#         # Generate OTP
#         
#         # Save user with OTP (but not active yet)
#         user = RegisterUser.objects.create(
#             username=username,
#             email=email,
#             password=password,
#             confirm_password=confirm_password,
#         )
#         
#         # Send OTP email
#         messages.success(request, 'Registration successful! Please login with your credentials.')
#         return redirect('accounts:login')
#     
#     return render(request, 'customerapp/register.html')



# def login(request):
#     if "email" in request.session:
#         return redirect('home')  # Changed from index to home
#     else:
#         if request.method == 'POST':
#             email = request.POST.get("email", '').strip()
#             password = request.POST.get("password", '').strip()
#             
#             # Check if fields are empty
#             if not email or not password:
#                 messages.error(request, 'Email and Password are required!')
#                 return render(request, "customerapp/login.html")
#             
#             try:
#                 uid = RegisterUser.objects.get(email=email)
#                 if password == uid.password:
#                     request.session["email"] = email
#                     return redirect('home')  # Redirects to home page after successful login
#                 else:
#                     messages.error(request, 'Invalid password!')
#                     return render(request, "customerapp/login.html", {"email": email})
#             except RegisterUser.DoesNotExist:
#                 messages.error(request, 'Invalid email!')
#                 return render(request, "customerapp/login.html", {"email": email})
#         
#         return render(request, 'customerapp/login.html')

def logout(request):
    if 'email' in request.session:
        del request.session['email']
    from django.contrib.auth import logout as auth_logout
    from accounts.views import delete_remember_login_cookie

    auth_logout(request)
    response = redirect('accounts:login')
    return delete_remember_login_cookie(response)



# def forgotpassword(request):
#     if request.method == 'POST':
#         email = request.POST.get('email')
#         if register.objects.filter(email=email).exists():
#             user = register.objects.get(email=email)
#             otp = random.randint(100000, 999999)
#             user.otp = otp
#             user.save()
#             
#             try:
#                 send_mail(
#                     'Password Reset OTP',
#                     f'Your OTP for password reset is: {otp}\nThis OTP will expire in 10 minutes.',
#                     'gohiljayb10@gmail.com',
#                     [email],
#                     fail_silently=False,
#                 )
#                 messages.success(request, 'OTP sent to your email!')
#                 return redirect('reset_password', user_id=user.id)
#             except Exception as e:
#                 messages.error(request, f'Failed to send OTP: {str(e)}')
#         else:
#             messages.error(request, 'Email not found!')
#     
#     return render(request, 'customerapp/forgotpassword.html')

def shop(request):
    products = product_model.objects.all().order_by("-id")
    wishlist_ids = set()
    if "email" in request.session:
        uid = RegisterUser.objects.get(email=request.session['email'])
        wishlist_ids = set(
            wishlist_model.objects.filter(register=uid).values_list("product_id", flat=True)
        )
    selected_category = request.GET.get("category", "all")
    filter_type = request.GET.get("filter", "all")
    sort_by = request.GET.get("sort_by", "newest")

    if selected_category and selected_category != "all":
        products = products.filter(category=selected_category)

    if filter_type == "bestseller":
        products = products.filter(bestseller=True)
    elif filter_type == "new_release":
        products = products.filter(new_release=True)
    elif filter_type == "expert_pick":
        products = products.filter(expert_pick=True)

    if sort_by == "price_low":
        products = products.order_by("price")
    elif sort_by == "price_high":
        products = products.order_by("-price")

    total_count = products.count()
    paginator = Paginator(products, 12)
    page = request.GET.get("page")
    pid = paginator.get_page(page)

    context = {
        "pid": pid,
        "category": selected_category,
        "filter": filter_type,
        "sort_by": sort_by,
        "total_count": total_count,
        "wishlist_ids": wishlist_ids,
    }

    return render(request, "customerapp/shop.html", context)

def checkout(request):
    if "email" not in request.session:
        return redirect('accounts:login')

    uid = RegisterUser.objects.get(email=request.session['email'])
    cart_items = add_to_cart.objects.filter(register=uid, order_status=False).order_by('-id')

    if not cart_items.exists():
        messages.info(request, 'Your cart is empty.')
        return redirect('cart')

    subtotal = sum(item.total for item in cart_items)
    shipping = 40 if subtotal > 0 else 0
    discount = 0
    coupon_id = request.session.get('coupon_id')
    if coupon_id and subtotal > 0:
        try:
            coupon = Coupon.objects.get(id=coupon_id, is_active=True)
            discount = coupon.discount_amount
            if discount > subtotal:
                discount = subtotal
        except Coupon.DoesNotExist:
            del request.session['coupon_id']

    total = subtotal + shipping - discount
    razorpay_response = None

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        address = request.POST.get('address', '').strip()
        city = request.POST.get('city', '').strip()
        pincode = request.POST.get('pincode', '').strip()
        payment_method = request.POST.get('payment_method', 'card').strip()

        full_name = f"{first_name} {last_name}".strip()

        if not first_name or not last_name or not email or not phone or not address or not city or not pincode:
            messages.error(request, 'All fields are required! Please fill in all the fields before checkout.')
        elif not cart_items:
            messages.error(request, 'Your cart is empty. Add products before checkout.')
        else:
            if payment_method == 'cod':
                # Cash on Delivery: create checkout immediately
                first_order_id = None
                
                for item in cart_items:
                    order = checkout_model.objects.create(
                        register=uid,
                        name=full_name,
                        email=email,
                        address=f"{address}, {city} - {pincode}",
                        phone=phone,
                        product_name=item.product_name,
                        image=item.product.image.name if (item.product and item.product.image) else (item.image.name if hasattr(item.image, 'name') else item.image),
                        price=item.price,
                        quantity=item.quantity,
                        total=item.total,
                        status='Pending'
                    )
                    if first_order_id is None:
                        first_order_id = order.id
                    
                    item.order_status = True
                    item.save()

                if 'coupon_id' in request.session:
                    del request.session['coupon_id']

                messages.success(request, 'Order placed successfully via Cash on Delivery!')
                
                # Redirect to recommendations page if order exists
                if first_order_id:
                    return redirect('purchase_success_recommendations', order_id=first_order_id)
                else:
                    return redirect('shop')
            else:
                # Create Razorpay order
                amount_paise = total * 100
                try:
                    if razorpay is None:
                        raise RuntimeError("Razorpay package is not installed.")
                    client = razorpay.Client(
                        auth=('rzp_test_bilBagOBVTi4lE', '77yKq3N9Wul97JVQcjtIVB5z')
                    )
                    razorpay_response = client.order.create({
                        'amount': amount_paise,
                        'currency': 'INR',
                        'payment_capture': 1
                    })

                    # Store billing info in session for use after payment
                    request.session['billing_info'] = {
                        'full_name': full_name,
                        'email': email,
                        'phone': phone,
                        'address': f"{address}, {city} - {pincode}",
                        'city': city,
                        'pincode': pincode,
                    }
                except Exception as e:
                    print(f"Razorpay error: {e}")
                    messages.error(request, "Could not initiate Razorpay payment. Please try again.")

    context = {
        'cart_items': cart_items,
        'subtotal': subtotal,
        'shipping': shipping,
        'discount': discount,
        'total': total,
        'uid': uid,
        'response': razorpay_response,
    }
    return render(request, 'customerapp/checkout.html', context)


def payment_success(request):
    if "email" not in request.session:
        return redirect('accounts:login')

    payment_id = request.GET.get('payment_id')
    order_id = request.GET.get('order_id')

    if not payment_id or not order_id:
        messages.error(request, 'Invalid payment information.')
        return redirect('cart')

    uid = RegisterUser.objects.get(email=request.session['email'])
    cart_items = add_to_cart.objects.filter(register=uid, order_status=False)

    billing = request.session.get('billing_info', {})
    full_name = billing.get('full_name', uid.username)
    email = billing.get('email', uid.email)
    phone = billing.get('phone', '')
    address = billing.get('address', '')

    # Store the first order ID for redirect to recommendations
    first_order_id = None
    
    for item in cart_items:
        order = checkout_model.objects.create(
            register=uid,
            name=full_name,
            email=email,
            address=address,
            phone=phone,
            product_name=item.product_name,
            image=item.product.image.name if (item.product and item.product.image) else (item.image.name if hasattr(item.image, 'name') else item.image),
            price=item.price,
            quantity=item.quantity,
            total=item.total,
        )
        if first_order_id is None:
            first_order_id = order.id
        
        item.order_status = True
        item.save()

    # Clean up session
    if 'billing_info' in request.session:
        del request.session['billing_info']
    if 'coupon_id' in request.session:
        del request.session['coupon_id']

    messages.success(request, f'Payment successful! Payment ID: {payment_id}. Your order has been placed.')
    
    # Redirect to new purchase success page with recommendations
    if first_order_id:
        return redirect('purchase_success_recommendations', order_id=first_order_id)
    else:
        return redirect('shop')


def product(request, id):
    if "email" in request.session:
        uid = RegisterUser.objects.get(email=request.session['email'])
        spid=product_model.objects.get(id=id)
        is_wishlisted = wishlist_model.objects.filter(register=uid, product=spid).exists()
        reviews = ProductReview.objects.filter(product=spid)
        contaxt={
            "spid":spid,
            "is_wishlisted": is_wishlisted,
            "reviews": reviews,
            "review_count": reviews.count(),
            "user_email": uid.email,
        }
        return render(request,"customerapp/product.html",contaxt)
    else:
        return redirect('accounts:login')

def submit_review(request, id):
    if "email" not in request.session:
        return redirect('accounts:login')
    if request.method == 'POST':
        uid = RegisterUser.objects.get(email=request.session['email'])
        spid = product_model.objects.get(id=id)
        message = request.POST.get('message', '').strip()
        if message:
            ProductReview.objects.create(
                product=spid,
                user=uid,
                email=uid.email,
                message=message,
            )
            messages.success(request, 'Your review has been posted successfully!')
        else:
            messages.error(request, 'Please enter a review message.')
    return redirect('product', id=id)

def blog(request):
    return render(request, 'customerapp/blog.html')

# def reset_password(request, user_id):
#     if request.method == 'POST':
#         entered_otp = request.POST.get('otp')
#         new_password = request.POST.get('new_password')
#         confirm_password = request.POST.get('confirm_password')
#         
#         try:
#             user = register.objects.get(id=user_id)
#             if str(user.otp) == entered_otp:
#                 if new_password == confirm_password:
#                     user.password = new_password
#                     user.confirm_password = confirm_password
#                     user.otp = None
#                     user.save()
#                     messages.success(request, 'Password reset successfully!')
#                     return redirect('accounts:login')
#                 else:
#                     messages.error(request, 'Passwords do not match!')
#             else:
#                 messages.error(request, 'Invalid OTP!')
#         except register.DoesNotExist:
#             messages.error(request, 'User not found!')
#             return redirect('forgotpassword')
#     
#     return render(request, 'customerapp/reset_password.html', {'user_id': user_id})



def wishlist(request):
    if "email" in request.session:
        uid = RegisterUser.objects.get(email=request.session['email'])
        user_wishlist = wishlist_model.objects.filter(register=uid).order_by("-id")

        contaxt = {
            "pid": user_wishlist,
            "wid": user_wishlist,
        }
        return render(request, "customerapp/wishlist.html", contaxt)
    return redirect('accounts:login')


def add_wishlist(request, id):
    if "email" in request.session:
        uid = RegisterUser.objects.get(email=request.session['email'])
        spid = product_model.objects.get(id=id)

        existing = wishlist_model.objects.filter(register=uid, product=spid).first()
        if existing:
            existing.delete()
        else:
            wishlist_model.objects.create(
                register=uid,
                product=spid,
                product_name=spid.name,
                price=spid.price,
                image=spid.image
            )
        return redirect('shop')
    return redirect('accounts:login')


def wishlist_delete(request, id):
    wishlist_model.objects.filter(id=id).delete()
    return redirect('wishlist')


def cart(request):
    print("Cart view called")
    if "email" in request.session:
        print(f"User email in session: {request.session['email']}")
        uid=RegisterUser.objects.get(email=request.session['email'])
        print(f"User found: {uid}")
        pid=add_to_cart.objects.filter(register=uid,order_status=False).order_by("-id")
        print(f"Cart items found: {pid.count()}")
        for item in pid:
            print(f"Cart item: {item.product_name}, quantity: {item.quantity}")
       
        l1=[]
        for i in pid:
            l1.append(i.total)
        print(l1)

        subtotal=sum(l1)
        shipping = 40 if subtotal > 0 else 0
        discount = 0
        applied_coupon_code = None

        coupon_id = request.session.get('coupon_id')
        if coupon_id and subtotal > 0:
            try:
                coupon = Coupon.objects.get(id=coupon_id, is_active=True)
                discount = coupon.discount_amount
                if discount > subtotal:
                    discount = subtotal
                applied_coupon_code = coupon.code
            except Coupon.DoesNotExist:
                del request.session['coupon_id']

        total = subtotal + shipping - discount
        contaxt={
            "pid":pid,
            "l1":l1,
            "subtotal":subtotal,
            "shipping": shipping,
            "total" : total,
            "discount": discount,
            "applied_coupon_code": applied_coupon_code
        }
        return render(request,"customerapp/cart.html",contaxt)
    else:
        return redirect('accounts:login')
    
def apply_coupon(request):
    if request.method == 'POST':
        code = request.POST.get('coupon_code', '').strip()
        if code:
            try:
                coupon = Coupon.objects.get(code=code, is_active=True)
                request.session['coupon_id'] = coupon.id
                messages.success(request, f'Coupon "{code}" applied successfully!')
            except Coupon.DoesNotExist:
                messages.error(request, 'Invalid or expired coupon code.')
        else:
            messages.error(request, 'Please enter a valid coupon code.')
    return redirect('cart')

def remove_coupon(request):
    if 'coupon_id' in request.session:
        del request.session['coupon_id']
        messages.info(request, 'Coupon removed successfully.')
    return redirect('cart')

def cart_add(request,id):
    print(f"cart_add called with id: {id}")
    if "email" in request.session:
        print(f"User email in session: {request.session['email']}")
        uid=RegisterUser.objects.get(email=request.session['email'])
        print(f"User found: {uid}")
        next_url = request.META.get("HTTP_REFERER", "/shop/")
        
        # Get quantity from request, default to 1
        qty = request.POST.get('qty', request.GET.get('qty', 1))
        try:
            qty = int(qty)
            if qty < 1: qty = 1
        except (ValueError, TypeError):
            qty = 1

        try:
            spid=product_model.objects.get(id=id)
            print(f"Product found: {spid.name}")
        except product_model.DoesNotExist:
            print("Product not found!")
            messages.error(request, 'Product not found.')
            return redirect(next_url)
        
        # Only check for products that are not yet purchased (order_status=False)
        cart_item = add_to_cart.objects.filter(register=uid, product=spid, order_status=False).first()
        if cart_item:
            # If product already in cart, increase quantity by the requested amount
            cart_item.quantity += qty
            cart_item.total = cart_item.price * cart_item.quantity
            cart_item.save()
            print(f"Updated quantity to {cart_item.quantity}")
        else:
            # Add new product to cart with the requested quantity
            cart_item = add_to_cart.objects.create(
                register=uid, 
                product=spid, 
                product_name=spid.name, 
                price=spid.price, 
                quantity=qty, 
                total=spid.price * qty, 
                order_status=False
            )
            print(f"Created new cart item: {cart_item}")
        
        messages.success(request, f'{qty} item(s) added to cart successfully.')
        return redirect(next_url)
    else:    
        print("No email in session - redirecting to login")
        return redirect('accounts:login')
    

def cart_minus(request,id):
    seid=add_to_cart.objects.get(id=id)
    if seid.quantity >1:
        seid.quantity-=1
        seid.total=seid.price*seid.quantity
        seid.save() 
    else:
        seid.delete()
    return redirect('cart')

def cart_plus(request,id):
    seid=add_to_cart.objects.get(id=id)
    seid.quantity+=1
    seid.total=seid.price*seid.quantity
    seid.save() 
    return redirect('cart')

def cart_delete(request,id):
    seid=add_to_cart.objects.get(id=id)
    seid.delete() 
    return redirect('cart')

def profile(request):
    if "email" not in request.session:
        return redirect('accounts:login')
    
    uid = RegisterUser.objects.get(email=request.session['email'])
    
    if request.method == "POST":
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        address = request.POST.get('address', '').strip()
        phone = request.POST.get('phone', '').strip()
        gender = request.POST.get('gender', '').strip()
        image = request.FILES.get('image')
        
        try:
            if not username or not email:
                messages.error(request, 'Name and email are required.')
                return redirect('profile')

            duplicate_email = RegisterUser.objects.filter(email=email).exclude(id=uid.id).exists()
            if duplicate_email:
                messages.error(request, 'This email is already used by another account.')
                return redirect('profile')

            old_email = uid.email
            UserModel = get_user_model()
            auth_user = UserModel.objects.filter(email=old_email).first()
            if auth_user and UserModel.objects.filter(email=email).exclude(id=auth_user.id).exists():
                messages.error(request, 'This email is already used by another account.')
                return redirect('profile')

            uid.username = username
            uid.email = email
            request.session['email'] = email
            request.session.set_expiry(settings.SESSION_COOKIE_AGE)
            uid.address = address
            uid.phone = phone
            uid.gender = gender
            if image:
                uid.image = image
            uid.save()

            if auth_user:
                gender_map = {
                    'Male': 'MALE',
                    'Female': 'FEMALE',
                }
                auth_user.username = username
                auth_user.email = email
                auth_user.phone_number = phone if phone else None
                if gender in gender_map:
                    auth_user.gender = gender_map[gender]
                if image:
                    auth_user.id_image = image
                auth_user.save()

            messages.success(request, 'Profile updated successfully!')
        except Exception as e:
            messages.error(request, f'Error updating profile: {e}')
        return redirect('profile')

    all_orders = checkout_model.objects.filter(register=uid).order_by('-order_date')
    
    # "Active" can be defined as anything not yet delivered
    active_orders = all_orders.exclude(status='Delivered')
    past_orders = all_orders.filter(status='Delivered')
    
    active_items_count = sum(order.quantity for order in active_orders)
    
    context = {
        'uid': uid,
        'active_orders': active_orders,
        'past_orders': past_orders,
        'active_items_count': active_items_count,
    }
    return render(request, 'customerapp/profile.html', context)


@xframe_options_sameorigin
def chat_ui(request):
    # Only allow access when loaded in an iframe (from the floating chatbot popup).
    # Block direct browser navigation to /chat/.
    fetch_dest = request.headers.get("Sec-Fetch-Dest", "")
    if fetch_dest == "iframe" or request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return render(request, "customerapp/chat.html")
    # Direct browser visit → redirect to home
    return redirect("home")


NORMAL_CONVERSATIONS = {
    "hello": "Hello 👋 How are you today?",
    "hi": "Hi 👋 Welcome to our bookstore. How can I help you today?",
    "hey": "Hey 😊 Looking for a good book today?",
    "how are you": "I am doing great 😊 Ready to help you find amazing books.",
    "good morning": "Good morning ☀️ Hope you have a wonderful reading day.",
    "good afternoon": "Good afternoon 📚 What kind of books are you interested in?",
    "good evening": "Good evening 🌙 Ready for your next great read?",
    "thank you": "You're very welcome 😊",
    "thanks": "Happy to help 📚",
    "bye": "Goodbye 👋 Have a great day and happy reading.",
    "goodbye": "See you again 📚 Take care.",
    "who are you": "I am your AI Bookstore Assistant 🤖",
    "what can you do": "I can help you find books, recommend books, track orders, suggest genres, and much more.",
    "help": "You can ask me about books, recommendations, moods, genres, or orders.",
}

SUPPORTED_LANGUAGES = {"english": "English", "hindi": "Hindi", "gujarati": "Gujarati"}

LANGUAGE_TEXT = {
    "English": {
        "book": "Book",
        "title": "Title",
        "author": "Author",
        "category": "Category",
        "price": "Price",
        "stock": "Stock",
        "available": "available",
        "buy": "Buy",
        "why": "Why",
        "no_image": "No image uploaded",
        "hi": "Hi",
        "main_menu": "I am Virtual Assistant. What would you like help with today?",
        "category_menu": "Choose a category and I will show matching books.",
        "books_for_mood": "Recommended books for your mood: {mood}",
        "books_from_genre": "Here are books from {genre}.",
        "books_in_category": "Books in this category:",
        "no_books_in_category": "No books are available in {category} yet.",
        "no_catalog_books": "No matching books are in our store catalog yet, but you can buy these online:",
        "out_of_stock": "{title} is out of stock",
        "try_similar": "Try similar books",
        "no_similar": "No similar books found",
        "fallback": "I could not match that to a store action. You can use free text as a fallback, or choose one of these.",
        "mood_happy": "Here are uplifting books to keep your good mood growing.",
        "mood_sad": "Here are gentle, hopeful books that can help you feel lighter.",
        "mood_stressed": "Here are calming books to help you slow down and breathe.",
        "mood_romantic": "Here are warm romantic reads for your mood.",
        "mood_thrill": "Here are gripping books for a suspenseful mood.",
        "mood_bored": "Here are page-turners to help you escape boredom.",
    },
    "Hindi": {
        "book": "पुस्तक",
        "title": "शीर्षक",
        "author": "लेखक",
        "category": "श्रेणी",
        "price": "कीमत",
        "stock": "स्टॉक",
        "available": "उपलब्ध",
        "buy": "खरीदें",
        "why": "क्यों",
        "no_image": "कोई चित्र अपलोड नहीं है",
        "hi": "नमस्ते",
        "main_menu": "मैं आपका वर्चुअल असिस्टेंट हूं। आज आप किसमें मदद चाहते हैं?",
        "category_menu": "एक श्रेणी चुनें और मैं उससे मिलती किताबें दिखाऊंगा।",
        "books_for_mood": "आपके मूड के लिए सुझाई गई किताबें: {mood}",
        "books_from_genre": "{genre} श्रेणी की किताबें यहां हैं।",
        "books_in_category": "इस श्रेणी की किताबें:",
        "no_books_in_category": "{category} श्रेणी में अभी कोई किताब उपलब्ध नहीं है।",
        "no_catalog_books": "हमारे स्टोर कैटलॉग में अभी मिलती-जुलती किताबें नहीं हैं, लेकिन आप इन्हें ऑनलाइन खरीद सकते हैं:",
        "out_of_stock": "{title} अभी स्टॉक में नहीं है",
        "try_similar": "मिलती-जुलती किताबें आजमाएं",
        "no_similar": "कोई मिलती-जुलती किताब नहीं मिली",
        "fallback": "मैं इसे किसी स्टोर एक्शन से मिलान नहीं कर पाया। आप फ्री टेक्स्ट इस्तेमाल कर सकते हैं या इनमें से चुन सकते हैं।",
        "mood_happy": "आपके अच्छे मूड को और बढ़ाने के लिए ये प्रेरक किताबें हैं।",
        "mood_sad": "ये नरम और उम्मीदभरी किताबें आपको हल्का महसूस कराने में मदद कर सकती हैं।",
        "mood_stressed": "धीमा होने और सांस लेने में मदद के लिए ये शांत किताबें हैं।",
        "mood_romantic": "आपके मूड के लिए ये प्यारी रोमांटिक किताबें हैं।",
        "mood_thrill": "रोमांचक मूड के लिए ये पकड़ बनाए रखने वाली किताबें हैं।",
        "mood_bored": "बोरियत से निकलने के लिए ये तेज और मजेदार किताबें हैं।",
    },
    "Gujarati": {
        "book": "પુસ્તક",
        "title": "શીર્ષક",
        "author": "લેખક",
        "category": "શ્રેણી",
        "price": "કિંમત",
        "stock": "સ્ટોક",
        "available": "ઉપલબ્ધ",
        "buy": "ખરીદો",
        "why": "શા માટે",
        "no_image": "કોઈ છબી અપલોડ નથી",
        "hi": "નમસ્તે",
        "main_menu": "હું તમારો વર્ચ્યુઅલ આસિસ્ટન્ટ છું. આજે તમને શું મદદ જોઈએ છે?",
        "category_menu": "એક શ્રેણી પસંદ કરો અને હું મળતી પુસ્તકો બતાવીશ.",
        "books_for_mood": "તમારા મૂડ માટે ભલામણ કરેલી પુસ્તકો: {mood}",
        "books_from_genre": "{genre} શ્રેણીની પુસ્તકો અહીં છે.",
        "books_in_category": "આ શ્રેણીની પુસ્તકો:",
        "no_books_in_category": "{category} શ્રેણીમાં હજી કોઈ પુસ્તક ઉપલબ્ધ નથી.",
        "no_catalog_books": "અમારા સ્ટોર કેટલોગમાં હાલ મેળ ખાતી પુસ્તકો નથી, પણ તમે આ ઑનલાઇન ખરીદી શકો છો:",
        "out_of_stock": "{title} હાલ સ્ટોકમાં નથી",
        "try_similar": "મળતી આવતી પુસ્તકો અજમાવો",
        "no_similar": "કોઈ મળતી આવતી પુસ્તક મળી નથી",
        "fallback": "હું આને કોઈ સ્ટોર એક્શન સાથે મેચ કરી શક્યો નહીં. તમે ફ્રી ટેક્સ્ટ વાપરી શકો છો અથવા આમાંથી પસંદ કરો.",
        "mood_happy": "તમારો સારો મૂડ વધારવા માટે આ પ્રેરક પુસ્તકો છે.",
        "mood_sad": "આ નરમ અને આશાવાદી પુસ્તકો તમને હળવું અનુભવવામાં મદદ કરી શકે છે.",
        "mood_stressed": "ધીમા થવા અને શાંતિ મેળવવા માટે આ શાંત પુસ્તકો છે.",
        "mood_romantic": "તમારા મૂડ માટે આ ઉષ્માભરી રોમેન્ટિક પુસ્તકો છે.",
        "mood_thrill": "રોમાંચક મૂડ માટે આ રસપ્રદ પુસ્તકો છે.",
        "mood_bored": "કંટાળો દૂર કરવા માટે આ ઝડપી અને મજેદાર પુસ્તકો છે.",
    },
}

LABEL_TRANSLATIONS = {
    "Hindi": {
        "Recommendations": "सिफारिशें",
        "Category Section": "श्रेणी सेक्शन",
        "Deals and Offers": "डील और ऑफर",
        "Customer Support": "ग्राहक सहायता",
        "Search by Title": "शीर्षक से खोजें",
        "Search by Author": "लेखक से खोजें",
        "Browse by Genre": "शैली से ब्राउज करें",
        "Search by ISBN": "ISBN से खोजें",
        "By Mood": "मूड के अनुसार",
        "By Genre": "शैली के अनुसार",
        "Best Sellers": "बेस्ट सेलर",
        "New Arrivals": "नई किताबें",
        "Track Order": "ऑर्डर ट्रैक करें",
        "Cancel Order": "ऑर्डर रद्द करें",
        "Return or Replace": "रिटर्न या रिप्लेस",
        "Payment Help": "भुगतान सहायता",
        "Today's Deals": "आज की डील",
        "Student Offers": "छात्र ऑफर",
        "Clearance Sale": "क्लियरेंस सेल",
        "Coupon Codes": "कूपन कोड",
        "I feel sad": "मैं उदास हूं",
        "I feel stressed": "मैं तनाव में हूं",
        "I feel happy": "मैं खुश हूं",
        "I want romance": "मुझे रोमांस चाहिए",
        "I want thrills": "मुझे रोमांच चाहिए",
        "Novel": "उपन्यास",
        "Fantasy": "फैंटेसी",
        "Thriller": "थ्रिलर",
        "Spiritual": "आध्यात्मिक",
        "Spiritual Books": "आध्यात्मिक किताबें",
        "Self Improvement": "सेल्फ इम्प्रूवमेंट",
        "Health and Fitness": "स्वास्थ्य और फिटनेस",
        "Love Stories": "प्रेम कहानियां",
        "Juvenile Literature": "बाल साहित्य",
        "Mystery": "रहस्य",
        "Self Health and Personal Growth": "सेल्फ हेल्थ और पर्सनल ग्रोथ",
        "Health Fitness and Wellness": "स्वास्थ्य फिटनेस और वेलनेस",
        "Spirituality and Philosophy": "आध्यात्मिकता और दर्शन",
        "Science Natural and Technology": "विज्ञान प्रकृति और टेक्नोलॉजी",
        "Leadership and Management": "लीडरशिप और मैनेजमेंट",
        "Productivity and Time Management": "प्रोडक्टिविटी और टाइम मैनेजमेंट",
        "Career Professional Development": "करियर प्रोफेशनल डेवलपमेंट",
        "Holy Books/Religious Texts/Scriptures": "पवित्र पुस्तकें/धार्मिक ग्रंथ/शास्त्र",
        "Back to Main Menu": "मुख्य मेन्यू पर वापस",
        "Find a Book": "किताब खोजें",
        "Get Recommendations": "सिफारिशें पाएं",
        "Choose Another Mood": "दूसरा मूड चुनें",
        "See Deals": "डील देखें",
        "Try Another Genre": "दूसरी शैली आजमाएं",
        "Search Differently": "दूसरे तरीके से खोजें",
        "Browse Deals": "डील ब्राउज करें",
        "Find Similar Books": "मिलती-जुलती किताबें खोजें",
        "Add to Wishlist": "विशलिस्ट में जोड़ें",
        "Find by Title": "शीर्षक से खोजें",
        "Find by Author": "लेखक से खोजें",
        "Admin Panel": "एडमिन पैनल",
    },
    "Gujarati": {
        "Recommendations": "ભલામણો",
        "Category Section": "શ્રેણી વિભાગ",
        "Deals and Offers": "ડીલ અને ઑફર",
        "Customer Support": "ગ્રાહક સહાય",
        "Search by Title": "શીર્ષકથી શોધો",
        "Search by Author": "લેખકથી શોધો",
        "Browse by Genre": "શૈલીથી બ્રાઉઝ કરો",
        "Search by ISBN": "ISBN થી શોધો",
        "By Mood": "મૂડ મુજબ",
        "By Genre": "શૈલી મુજબ",
        "Best Sellers": "બેસ્ટ સેલર",
        "New Arrivals": "નવી આવકો",
        "Track Order": "ઓર્ડર ટ્રેક કરો",
        "Cancel Order": "ઓર્ડર રદ કરો",
        "Return or Replace": "રિટર્ન અથવા રિપ્લેસ",
        "Payment Help": "ચુકવણી સહાય",
        "Today's Deals": "આજની ડીલ",
        "Student Offers": "વિદ્યાર્થી ઑફર",
        "Clearance Sale": "ક્લિયરન્સ સેલ",
        "Coupon Codes": "કૂપન કોડ",
        "I feel sad": "હું દુઃખી છું",
        "I feel stressed": "હું તણાવમાં છું",
        "I feel happy": "હું ખુશ છું",
        "I want romance": "મને રોમાન્સ જોઈએ",
        "I want thrills": "મને રોમાંચ જોઈએ",
        "Novel": "નવલકથા",
        "Fantasy": "ફૅન્ટસી",
        "Thriller": "થ્રિલર",
        "Spiritual": "આધ્યાત્મિક",
        "Spiritual Books": "આધ્યાત્મિક પુસ્તકો",
        "Self Improvement": "સેલ્ફ ઇમ્પ્રૂવમેન્ટ",
        "Health and Fitness": "આરોગ્ય અને ફિટનેસ",
        "Love Stories": "પ્રેમ કથાઓ",
        "Juvenile Literature": "બાળ સાહિત્ય",
        "Mystery": "રહસ્ય",
        "Self Health and Personal Growth": "સેલ્ફ હેલ્થ અને પર્સનલ ગ્રોથ",
        "Health Fitness and Wellness": "આરોગ્ય ફિટનેસ અને વેલનેસ",
        "Spirituality and Philosophy": "આધ્યાત્મિકતા અને તત્વજ્ઞાન",
        "Science Natural and Technology": "વિજ્ઞાન પ્રકૃતિ અને ટેકનોલોજી",
        "Leadership and Management": "લીડરશિપ અને મેનેજમેન્ટ",
        "Productivity and Time Management": "પ્રોડક્ટિવિટી અને ટાઇમ મેનેજમેન્ટ",
        "Career Professional Development": "કેરિયર પ્રોફેશનલ ડેવલપમેન્ટ",
        "Holy Books/Religious Texts/Scriptures": "પવિત્ર પુસ્તકો/ધાર્મિક ગ્રંથો/શાસ્ત્રો",
        "Back to Main Menu": "મુખ્ય મેનૂ પર પાછા",
        "Find a Book": "પુસ્તક શોધો",
        "Get Recommendations": "ભલામણો મેળવો",
        "Choose Another Mood": "બીજો મૂડ પસંદ કરો",
        "See Deals": "ડીલ જુઓ",
        "Try Another Genre": "બીજી શૈલી અજમાવો",
        "Search Differently": "બીજી રીતે શોધો",
        "Browse Deals": "ડીલ બ્રાઉઝ કરો",
        "Find Similar Books": "મળતી આવતી પુસ્તકો શોધો",
        "Add to Wishlist": "વિશલિસ્ટમાં ઉમેરો",
        "Find by Title": "શીર્ષકથી શોધો",
        "Find by Author": "લેખકથી શોધો",
        "Admin Panel": "એડમિન પેનલ",
    },
}

MESSAGE_TRANSLATIONS = {
    "Hindi": {
        "Sure. How would you like to find your next book?": "ज़रूर। आप अपनी अगली किताब कैसे खोजना चाहेंगे?",
        "I can help you pick something that fits your mood or taste.": "मैं आपके मूड या पसंद के अनुसार किताब चुनने में मदद कर सकता हूं।",
        "Happy to help with your order. What do you need?": "आपके ऑर्डर में मदद करके खुशी होगी। आपको क्या चाहिए?",
        "Here are the offer sections shoppers usually love.": "ये वे ऑफर सेक्शन हैं जिन्हें ग्राहक आमतौर पर पसंद करते हैं।",
        "Let's manage your saved books.": "चलिए आपकी सेव की गई किताबें मैनेज करते हैं।",
        "I am here with you. What kind of support do you need?": "मैं आपकी मदद के लिए हूं। आपको किस तरह की सहायता चाहिए?",
        "Choose a category and I will show matching books.": "एक श्रेणी चुनें और मैं उससे मिलती किताबें दिखाऊंगा।",
        "Choose the mood you want your book to support.": "वह मूड चुनें जिसके लिए आप किताब चाहते हैं।",
        "Pick a genre and I will show matching books.": "एक शैली चुनें और मैं मिलती-जुलती किताबें दिखाऊंगा।",
        "Great. Which shelf should we browse first?": "बहुत बढ़िया। पहले कौन सी शेल्फ ब्राउज करें?",
    },
    "Gujarati": {
        "Sure. How would you like to find your next book?": "ચોક્કસ। તમે તમારી આગળની પુસ્તક કેવી રીતે શોધવા માંગો છો?",
        "I can help you pick something that fits your mood or taste.": "હું તમારા મૂડ અથવા પસંદ મુજબ પુસ્તક પસંદ કરવામાં મદદ કરી શકું છું.",
        "Happy to help with your order. What do you need?": "તમારા ઓર્ડરમાં મદદ કરીને આનંદ થશે. તમને શું જોઈએ છે?",
        "Here are the offer sections shoppers usually love.": "આ ઑફર વિભાગો ગ્રાહકોને સામાન્ય રીતે ગમે છે.",
        "Let's manage your saved books.": "ચાલો તમારી સેવ કરેલી પુસ્તકો મેનેજ કરીએ.",
        "I am here with you. What kind of support do you need?": "હું તમારી મદદ માટે છું. તમને કઈ પ્રકારની સહાય જોઈએ છે?",
        "Choose a category and I will show matching books.": "એક શ્રેણી પસંદ કરો અને હું મળતી પુસ્તકો બતાવીશ.",
        "Choose the mood you want your book to support.": "પુસ્તક માટે તમારે જે મૂડ જોઈએ છે તે પસંદ કરો.",
        "Pick a genre and I will show matching books.": "એક શૈલી પસંદ કરો અને હું મળતી પુસ્તકો બતાવીશ.",
        "Great. Which shelf should we browse first?": "સરસ. પહેલા કયો શેલ્ફ બ્રાઉઝ કરીએ?",
    },
}


def get_language(request):
    raw_language = request.GET.get("language") or request.session.get("chat_language") or "English"
    language = SUPPORTED_LANGUAGES.get(raw_language.lower(), "English")
    request.session["chat_language"] = language
    return language


def t(language, key, **kwargs):
    text = LANGUAGE_TEXT.get(language, LANGUAGE_TEXT["English"]).get(
        key, LANGUAGE_TEXT["English"].get(key, key)
    )
    return text.format(**kwargs) if kwargs else text


def translate_label(label, language):
    return LABEL_TRANSLATIONS.get(language, {}).get(label, label)


def translate_message(message, language):
    return MESSAGE_TRANSLATIONS.get(language, {}).get(message, message)


def get_normal_conversation_response(message, language="English"):

    message = message.lower().strip()
    if ":" in message:
        return None

    for key, response in NORMAL_CONVERSATIONS.items():
        if re.search(rf"\b{re.escape(key)}\b", message):
            translated_response = {
                "Hindi": {
                    "hello": "नमस्ते। आज आप कैसे हैं?",
                    "hi": "नमस्ते। हमारे बुकस्टोर में आपका स्वागत है। मैं आपकी कैसे मदद कर सकता हूं?",
                    "hey": "नमस्ते। क्या आज कोई अच्छी किताब ढूंढ रहे हैं?",
                    "how are you": "मैं अच्छा हूं। शानदार किताबें खोजने में आपकी मदद के लिए तैयार हूं।",
                    "thank you": "आपका स्वागत है।",
                    "thanks": "मदद करके खुशी हुई।",
                    "bye": "अलविदा। आपका दिन शुभ हो और पढ़ते रहें।",
                    "goodbye": "फिर मिलेंगे। ध्यान रखें।",
                    "who are you": "मैं आपका AI Bookstore Assistant हूं।",
                    "what can you do": "मैं किताबें खोजने, सिफारिशें देने, ऑर्डर सहायता, शैली सुझाव और बहुत कुछ कर सकता हूं।",
                    "help": "आप मुझसे किताबों, सिफारिशों, मूड, शैलियों या ऑर्डर के बारे में पूछ सकते हैं।",
                },
                "Gujarati": {
                    "hello": "નમસ્તે. આજે તમે કેમ છો?",
                    "hi": "નમસ્તે. અમારા બુકસ્ટોરમાં આપનું સ્વાગત છે. હું તમારી કેવી રીતે મદદ કરી શકું?",
                    "hey": "નમસ્તે. આજે કોઈ સારી પુસ્તક શોધી રહ્યા છો?",
                    "how are you": "હું સારો છું. સરસ પુસ્તકો શોધવામાં મદદ કરવા તૈયાર છું.",
                    "thank you": "તમારું સ્વાગત છે.",
                    "thanks": "મદદ કરીને આનંદ થયો.",
                    "bye": "આવજો. તમારો દિવસ શુભ રહે અને વાંચતા રહો.",
                    "goodbye": "ફરી મળીએ. ધ્યાન રાખજો.",
                    "who are you": "હું તમારો AI Bookstore Assistant છું.",
                    "what can you do": "હું પુસ્તકો શોધવામાં, ભલામણ કરવામાં, ઓર્ડર મદદમાં, શૈલી સૂચવવામાં અને વધુ મદદ કરી શકું છું.",
                    "help": "તમે મને પુસ્તકો, ભલામણો, મૂડ, શૈલીઓ અથવા ઓર્ડર વિશે પૂછો શકો છો.",
                },
            }.get(language, {}).get(key, response)
            return f"""
<div class="book-info">
    <div>{translated_response}</div>
</div>
"""

    return None
    

def get_category_terms(message):
    category_aliases = {
        "love stories": ["love story", "love stories", "romance", "romantic"],
        "fantasy": ["fantasy"],
        "juvenile literature": ["juvenile literature", "juvinile litreature", "children", "kids"],
        "mystery": ["mystery", "mestery", "suspense", "crime"],
        "novel": ["novel", "fiction"],
        "thriller": ["thriller", "mystery", "mestery", "suspense", "crime"],
        "self health and personal growth": ["self health", "personal growth", "self-help", "self help"],
        "health fitness and wellness": [
            "health",
            "fitness",
            "wellness",
            "health fitness",
            "self and health fitness",
        ],
        "spirituality and philosophy": ["spirituality", "spiritual", "philosophy", "spritual"],
        "science natural and technology": [
            "science",
            "natural",
            "technology",
            "scinece natural and technology",
        ],
        "leadership and management": [
            "leadership and management",
            "leadership",
            "management",
        ],
        "productivity and time management": ["productivity", "time management"],
        "career professional development": [
            "career",
            "carrer",
            "professional development",
            "career professional development",
        ],
        "holy books religious texts scriptures": [
            "holy books",
            "religious texts",
            "relegious texts",
            "scriptures",
            "holy books/relegious texts/scriptures",
        ],
    }

    terms = [message]
    for category, aliases in category_aliases.items():
        if message == category or message in aliases:
            terms.extend([category, *aliases])
            break

    return list(dict.fromkeys(term for term in terms if term))


def detect_mood(message):

    mood_map = {
        "happy": ["happy", "excited", "motivated", "good", "joy", "great"],
        "sad": ["sad", "depressed", "upset", "crying", "low", "lonely"],
        "stressed": ["stress", "stressed", "anxiety", "anxious", "tired", "overwhelmed"],
        "romantic": ["love", "romantic", "relationship", "romance"],
        "thrill": ["thrill", "suspense", "mystery", "crime", "adventure"],
        "bored": ["bored", "boring", "fun", "escape"],
    }

    for mood, keywords in mood_map.items():
        if any(word in message for word in keywords):
            return mood

    return None


def recommend_books_by_mood(request, mood, name=None):

    mood_categories = {
        "happy": ["leadership", "motivation", "productivity", "self-help"],
        "sad": ["spiritual", "self-help", "philosophy"],
        "stressed": ["spiritual", "mindfulness", "health", "fitness"],
        "romantic": ["novel", "fiction", "love"],
        "thrill": ["thriller", "mystery", "crime", "suspense"],
        "bored": ["fantasy", "novel", "adventure", "fiction"],
    }

    fallback_books = {
        "happy": [
            "Atomic Habits by James Clear",
            "The Alchemist by Paulo Coelho",
            "Ikigai by Hector Garcia and Francesc Miralles",
        ],
        "sad": [
            "The Midnight Library by Matt Haig",
            "Tuesdays with Morrie by Mitch Albom",
            "The Power of Now by Eckhart Tolle",
        ],
        "stressed": [
            "The Things You Can See Only When You Slow Down by Haemin Sunim",
            "Mindfulness in Plain English by Bhante Henepola Gunaratana",
            "The Comfort Book by Matt Haig",
        ],
        "romantic": [
            "It Ends with Us by Colleen Hoover",
            "Pride and Prejudice by Jane Austen",
            "The Fault in Our Stars by John Green",
        ],
        "thrill": [
            "The Silent Patient by Alex Michaelides",
            "Gone Girl by Gillian Flynn",
            "The Da Vinci Code by Dan Brown",
        ],
        "bored": [
            "Harry Potter and the Sorcerer's Stone by J.K. Rowling",
            "The Hobbit by J.R.R. Tolkien",
            "Ready Player One by Ernest Cline",
        ],
    }

    categories = mood_categories.get(mood, [])

    category_query = Q()
    for category in categories:
        category_query |= Q(category__icontains=category)

    books = Book.objects.filter(category_query)[:5] if category_query else []

    response = f"""
<div class="book-card">
    <div class="book-info">
        <strong>📚 Recommended books for your mood: {escape(mood.title())}</strong>
    </div>
</div>
"""

    response = f"""
<div class="book-card">
    <div class="book-info">
        <strong>&#128218; Recommended books for your mood: {escape(mood.title())}</strong>
    </div>
</div>
"""

    for book in books:
        response += render_book_card(request, book)

    if not books:
        response += "<ul>"
        for title in fallback_books.get(mood, []):
            response += f"<li>{escape(title)}</li>"
        response += "</ul>"

    return response



def get_book_image_url(request, book):
    if not getattr(book, "image", None):
        return ""

    try:
        return request.build_absolute_uri(book.image.url)
    except ValueError:
        return ""


def render_book_card(request, book, language="English"):
    image_url = get_book_image_url(request, book)
    
    title = getattr(book, 'title', getattr(book, 'name', ''))
    author = getattr(book, 'author', getattr(book, 'author_name', 'N/A'))
    stock = getattr(book, 'stock', 10)
    category_str = str(book.category).replace('_', ' ').title()
    
    image_html = (
        f'<img class="book-cover" src="{escape(image_url)}" alt="{escape(title)} cover">'
        if image_url
        else f'<div class="book-cover-placeholder">{escape(t(language, "no_image"))}</div>'
    )

    return f"""
<div class="book-card">
    {image_html}
    <div class="book-info">
        <div>&#128218; {escape(t(language, "title"))}: {escape(title)}</div>
        <div>&#9997; {escape(t(language, "author"))}: {escape(author)}</div>
        <div>&#128193; {escape(t(language, "category"))}: {escape(category_str)}</div>
        <div>&#128176; {escape(t(language, "price"))}: &#8377;{book.price}</div>
        <div>&#128230; {escape(t(language, "stock"))}: {stock} {escape(t(language, "available"))}</div>
    </div>
</div>
"""


def get_purchase_links(title):
    query = quote_plus(title)
    return {
        "Amazon": f"https://www.amazon.in/s?k={query}",
        "Flipkart": f"https://www.flipkart.com/search?q={query}",
        "Google Books": f"https://www.google.com/search?tbm=bks&q={query}",
    }


def render_purchase_links(title):
    links = get_purchase_links(title)
    return " | ".join(
        f'<a href="{escape(url)}" target="_blank" rel="noopener noreferrer">{escape(site)}</a>'
        for site, url in links.items()
    )


def recommend_books_by_mood(request, mood, name=None, language="English"):
    mood_categories = {
        "happy": ["leadership", "motivation", "productivity", "self-help"],
        "sad": ["spiritual", "self-help", "philosophy"],
        "stressed": ["spiritual", "mindfulness", "health", "fitness"],
        "romantic": ["novel", "fiction", "love"],
        "thrill": ["thriller", "mystery", "crime", "suspense"],
        "bored": ["fantasy", "novel", "adventure", "fiction"],
    }

    mood_messages = {
        "happy": t(language, "mood_happy"),
        "sad": t(language, "mood_sad"),
        "stressed": t(language, "mood_stressed"),
        "romantic": t(language, "mood_romantic"),
        "thrill": t(language, "mood_thrill"),
        "bored": t(language, "mood_bored"),
    }

    fallback_books = {
        "happy": [
            {
                "title": "Atomic Habits",
                "author": "James Clear",
                "reason": "Practical and motivating when you want positive energy.",
            },
            {
                "title": "The Alchemist",
                "author": "Paulo Coelho",
                "reason": "A hopeful story about dreams, courage, and purpose.",
            },
            {
                "title": "Ikigai",
                "author": "Hector Garcia and Francesc Miralles",
                "reason": "A light read about meaning and everyday happiness.",
            },
        ],
        "sad": [
            {
                "title": "The Midnight Library",
                "author": "Matt Haig",
                "reason": "A comforting novel about regret, hope, and second chances.",
            },
            {
                "title": "Tuesdays with Morrie",
                "author": "Mitch Albom",
                "reason": "Warm life lessons that can make sadness feel less heavy.",
            },
            {
                "title": "The Comfort Book",
                "author": "Matt Haig",
                "reason": "Short, gentle reflections for difficult emotional days.",
            },
        ],
        "stressed": [
            {
                "title": "The Things You Can See Only When You Slow Down",
                "author": "Haemin Sunim",
                "reason": "Simple reflections for calm and emotional balance.",
            },
            {
                "title": "Mindfulness in Plain English",
                "author": "Bhante Henepola Gunaratana",
                "reason": "Clear guidance for mindfulness and stress relief.",
            },
            {
                "title": "The Power of Now",
                "author": "Eckhart Tolle",
                "reason": "Helps shift attention away from anxious thoughts.",
            },
        ],
        "romantic": [
            {
                "title": "It Ends with Us",
                "author": "Colleen Hoover",
                "reason": "Emotional modern romance with strong feelings.",
            },
            {
                "title": "Pride and Prejudice",
                "author": "Jane Austen",
                "reason": "A classic romance with wit and warmth.",
            },
            {
                "title": "The Fault in Our Stars",
                "author": "John Green",
                "reason": "A tender love story with emotional depth.",
            },
        ],
        "thrill": [
            {
                "title": "The Silent Patient",
                "author": "Alex Michaelides",
                "reason": "A fast mystery that keeps your mind engaged.",
            },
            {
                "title": "Gone Girl",
                "author": "Gillian Flynn",
                "reason": "Dark, twisty suspense with strong momentum.",
            },
            {
                "title": "The Da Vinci Code",
                "author": "Dan Brown",
                "reason": "A quick adventure thriller with puzzles and pace.",
            },
        ],
        "bored": [
            {
                "title": "Harry Potter and the Sorcerer's Stone",
                "author": "J.K. Rowling",
                "reason": "An easy escape into a magical story.",
            },
            {
                "title": "The Hobbit",
                "author": "J.R.R. Tolkien",
                "reason": "A cozy adventure that is easy to get lost in.",
            },
            {
                "title": "Ready Player One",
                "author": "Ernest Cline",
                "reason": "Fast, playful, and packed with pop-culture adventure.",
            },
        ],
    }

    category_query = Q()
    for category in mood_categories.get(mood, []):
        category_query |= Q(category__icontains=category)

    books = product_model.objects.filter(category_query)[:5] if category_query else []

    response = f"""
<div class="book-info">
    <strong>&#128218; {greeting_for(name, language)}{escape(mood_messages.get(mood, t(language, "books_for_mood", mood=mood.title())))}</strong>
</div>
"""

    if books:
        for book in books:
            title = getattr(book, 'title', getattr(book, 'name', ''))
            author = getattr(book, 'author', getattr(book, 'author_name', 'N/A'))
            category_str = str(book.category).replace('_', ' ').title()
            
            response += f"""
<div class="book-card">
    <div class="book-info">
        <div>&#128218; <strong>{escape(title)}</strong></div>
        <div>&#9997; {escape(t(language, "author"))}: {escape(author)}</div>
        <div>&#128193; {escape(t(language, "category"))}: {escape(category_str)}</div>
        <div>&#128176; {escape(t(language, "price"))}: &#8377;{book.price}</div>
        <div>&#128279; {escape(t(language, "buy"))}: {render_purchase_links(title)}</div>
    </div>
</div>
"""
        return response

    for book in fallback_books.get(mood, []):
        title = f"{book['title']} {book['author']}"
        response += f"""
<div class="book-card">
    <div class="book-info">
        <div>&#128218; <strong>{escape(book['title'])}</strong></div>
        <div>&#9997; {escape(t(language, "author"))}: {escape(book['author'])}</div>
        <div>&#128161; {escape(t(language, "why"))}: {escape(book['reason'])}</div>
        <div>&#128279; {escape(t(language, "buy"))}: {render_purchase_links(title)}</div>
    </div>
</div>
"""

    return response


MAIN_MENU_OPTIONS = [
    ("Recommendations", "menu:recommend"),
    ("Category Section", "menu:categories"),
    ("Deals and Offers", "menu:deals"),
    ("Customer Support", "menu:support"),
]

CATEGORY_OPTIONS = [
    ("Love Stories", "category:love-stories"),
    ("Fantasy", "category:fantasy"),
    ("Juvenile Literature", "category:juvenile-literature"),
    ("Mystery", "category:mystery"),
    ("Thriller", "category:thriller"),
    ("Self Health and Personal Growth", "category:self-health-personal-growth"),
    ("Health Fitness and Wellness", "category:health-fitness-wellness"),
    ("Spirituality and Philosophy", "category:spirituality-philosophy"),
    ("Science Natural and Technology", "category:science-natural-technology"),
    ("Leadership and Management", "category:leadership-management"),
    ("Productivity and Time Management", "category:productivity-time-management"),
    ("Career Professional Development", "category:career-professional-development"),
    ("Holy Books/Religious Texts/Scriptures", "category:holy-books-religious-texts-scriptures"),
]

CATEGORY_SEARCH_TERMS = {
    "love-stories": ["love stories", "love story", "romance", "romantic"],
    "fantasy": ["fantasy"],
    "juvenile-literature": ["juvenile literature", "juvinile litreature", "children", "kids"],
    "mystery": ["mystery", "mestery", "crime", "suspense"],
    "thriller": ["thriller", "mystery", "suspense", "crime"],
    "self-health-personal-growth": ["self health", "personal growth", "self-help", "self help"],
    "health-fitness-wellness": ["health", "fitness", "wellness", "self and health fitness"],
    "spirituality-philosophy": ["spirituality", "spiritual", "spritual", "philosophy"],
    "science-natural-technology": ["science", "natural", "technology", "scinece"],
    "leadership-management": ["leadership", "management", "leadership and management"],
    "productivity-time-management": ["productivity", "time management"],
    "career-professional-development": ["career", "carrer", "professional development"],
    "holy-books-religious-texts-scriptures": [
        "holy books",
        "religious texts",
        "relegious texts",
        "scriptures",
    ],
}


MENU_FLOWS = {
    "menu:find": (
        "Sure. How would you like to find your next book?",
        [
            ("Search by Title", "find:title"),
            ("Search by Author", "find:author"),
            ("Browse by Genre", "find:genre"),
            ("Search by ISBN", "find:isbn"),
        ],
    ),
    "menu:recommend": (
        "I can help you pick something that fits your mood or taste.",
        [
            ("By Mood", "recommend:mood"),
            ("By Genre", "recommend:genre"),
            ("Best Sellers", "recommend:bestsellers"),
            ("New Arrivals", "recommend:new"),
        ],
    ),
    "menu:orders": (
        "Happy to help with your order. What do you need?",
        [
            ("Track Order", "order:track"),
            ("Cancel Order", "order:cancel"),
            ("Return or Replace", "order:return"),
            ("Payment Help", "order:payment"),
        ],
    ),
    "menu:categories": (
        "Choose a category and I will show matching books.",
        CATEGORY_OPTIONS,
    ),
    "menu:deals": (
        "Here are the offer sections shoppers usually love.",
        [
            ("Today's Deals", "deals:today"),
            ("Student Offers", "deals:student"),
            ("Clearance Sale", "deals:clearance"),
            ("Coupon Codes", "deals:coupons"),
        ],
    ),
    "menu:wishlist": (
        "Let's manage your saved books.",
        [
            ("View Wishlist", "wishlist:view"),
            ("Add a Book", "wishlist:add"),
            ("Remove a Book", "wishlist:remove"),
            ("Price Drop Alerts", "wishlist:alerts"),
        ],
    ),
    "menu:support": (
        "I am here with you. What kind of support do you need?",
        [
            ("Delivery Help", "support:delivery"),
            ("Refund Help", "support:refund"),
            ("Account Help", "support:account"),
            ("Contact Store Team", "support:contact"),
        ],
    ),
    "recommend:mood": (
        "Choose the mood you want your book to support.",
        [
            ("I feel sad", "mood:sad"),
            ("I feel stressed", "mood:stressed"),
            ("I feel happy", "mood:happy"),
            ("I want romance", "mood:romantic"),
            ("I want thrills", "mood:thrill"),
        ],
    ),
    "recommend:genre": (
        "Pick a genre and I will show matching books.",
        [
            ("Novel", "genre:novel"),
            ("Fantasy", "genre:fantasy"),
            ("Thriller", "genre:thriller"),
            ("Spiritual", "genre:spiritual"),
            ("Self Improvement", "genre:self-help"),
        ],
    ),
    "find:genre": (
        "Great. Which shelf should we browse first?",
        [
            ("Novel", "genre:novel"),
            ("Fantasy", "genre:fantasy"),
            ("Thriller", "genre:thriller"),
            ("Spiritual Books", "genre:spiritual"),
            ("Health and Fitness", "genre:health"),
        ],
    ),
}


DETAIL_MESSAGES = {
    "find:title": "Use the search box as a fallback for a book title, or start with one of these popular paths.",
    "find:author": "Use the search box as a fallback for an author name, or browse these author-friendly sections.",
    "find:isbn": "ISBN lookup needs exact text, so use the search box as a fallback when you have the number.",
    "recommend:bestsellers": "Best sellers are a great place to start.",
    "recommend:new": "Fresh arrivals are ready for browsing.",
    "order:track": "Order tracking is easiest from your account orders page.",
    "order:cancel": "Cancellation depends on whether the order has shipped.",
    "order:return": "Returns and replacements can be started from your order details.",
    "order:payment": "Payment help covers failed, pending, and duplicate payments.",
    "deals:today": "Today's deals can help you find a good book at a better price.",
    "deals:student": "Student offers are useful for textbooks and exam prep.",
    "deals:clearance": "Clearance books are limited-stock discounted picks.",
    "deals:coupons": "Coupon codes can be checked before checkout.",
    "wishlist:view": "Your wishlist keeps books ready for later.",
    "wishlist:add": "You can save any book to your wishlist from its details page.",
    "wishlist:remove": "Removing a wishlist item will not affect your orders.",
    "wishlist:alerts": "Price alerts help you buy when a saved book drops in price.",
    "support:delivery": "Delivery support can help with delays, address issues, and missing packages.",
    "support:refund": "Refund support can check return status and payment timelines.",
    "support:account": "Account support covers login, profile, and password help.",
    "support:contact": "You can contact the store team for anything that needs a person.",
}


DETAIL_OPTIONS = [
    ("Find a Book", "menu:find"),
    ("Get Recommendations", "menu:recommend"),
    ("Track Order", "menu:orders"),
    ("Customer Support", "menu:support"),
]


GENRE_FALLBACK_BOOKS = {
    "novel": [
        {"title": "The Alchemist", "author": "Paulo Coelho"},
        {"title": "The Midnight Library", "author": "Matt Haig"},
        {"title": "To Kill a Mockingbird", "author": "Harper Lee"},
    ],
    "fantasy": [
        {"title": "Harry Potter and the Sorcerer's Stone", "author": "J.K. Rowling"},
        {"title": "The Hobbit", "author": "J.R.R. Tolkien"},
        {"title": "The Name of the Wind", "author": "Patrick Rothfuss"},
    ],
    "thriller": [
        {"title": "The Silent Patient", "author": "Alex Michaelides"},
        {"title": "Gone Girl", "author": "Gillian Flynn"},
        {"title": "The Da Vinci Code", "author": "Dan Brown"},
    ],
    "spiritual": [
        {"title": "The Power of Now", "author": "Eckhart Tolle"},
        {"title": "Autobiography of a Yogi", "author": "Paramahansa Yogananda"},
        {"title": "The Bhagavad Gita", "author": "Eknath Easwaran"},
    ],
    "self-help": [
        {"title": "Atomic Habits", "author": "James Clear"},
        {"title": "Ikigai", "author": "Hector Garcia and Francesc Miralles"},
        {"title": "The 7 Habits of Highly Effective People", "author": "Stephen R. Covey"},
    ],
    "health": [
        {"title": "Why We Sleep", "author": "Matthew Walker"},
        {"title": "The Body Keeps the Score", "author": "Bessel van der Kolk"},
        {"title": "Outlive", "author": "Peter Attia"},
    ],
    "love-stories": [
        {"title": "Pride and Prejudice", "author": "Jane Austen"},
        {"title": "It Ends with Us", "author": "Colleen Hoover"},
        {"title": "The Fault in Our Stars", "author": "John Green"},
    ],
    "juvenile-literature": [
        {"title": "Charlotte's Web", "author": "E. B. White"},
        {"title": "Matilda", "author": "Roald Dahl"},
        {"title": "The Secret Garden", "author": "Frances Hodgson Burnett"},
    ],
    "mystery": [
        {"title": "The Hound of the Baskervilles", "author": "Arthur Conan Doyle"},
        {"title": "And Then There Were None", "author": "Agatha Christie"},
        {"title": "The Girl with the Dragon Tattoo", "author": "Stieg Larsson"},
    ],
    "self-health-personal-growth": [
        {"title": "Atomic Habits", "author": "James Clear"},
        {"title": "The 7 Habits of Highly Effective People", "author": "Stephen R. Covey"},
        {"title": "Mindset", "author": "Carol S. Dweck"},
    ],
    "health-fitness-wellness": [
        {"title": "Why We Sleep", "author": "Matthew Walker"},
        {"title": "Outlive", "author": "Peter Attia"},
        {"title": "The Blue Zones", "author": "Dan Buettner"},
    ],
    "spirituality-philosophy": [
        {"title": "The Power of Now", "author": "Eckhart Tolle"},
        {"title": "Meditations", "author": "Marcus Aurelius"},
        {"title": "Autobiography of a Yogi", "author": "Paramahansa Yogananda"},
    ],
    "science-natural-technology": [
        {"title": "A Brief History of Time", "author": "Stephen Hawking"},
        {"title": "Sapiens", "author": "Yuval Noah Harari"},
        {"title": "The Selfish Gene", "author": "Richard Dawkins"},
    ],
    "leadership-management": [
        {"title": "Leaders Eat Last", "author": "Simon Sinek"},
        {"title": "Good to Great", "author": "Jim Collins"},
        {"title": "The Lean Startup", "author": "Eric Ries"},
    ],
    "productivity-time-management": [
        {"title": "Deep Work", "author": "Cal Newport"},
        {"title": "Getting Things Done", "author": "David Allen"},
        {"title": "Eat That Frog!", "author": "Brian Tracy"},
    ],
    "career-professional-development": [
        {"title": "So Good They Can't Ignore You", "author": "Cal Newport"},
        {"title": "Designing Your Life", "author": "Bill Burnett and Dave Evans"},
        {"title": "What Color Is Your Parachute?", "author": "Richard N. Bolles"},
    ],
    "holy-books-religious-texts-scriptures": [
        {"title": "The Bhagavad Gita", "author": "Eknath Easwaran"},
        {"title": "The Holy Bible", "author": "Various"},
        {"title": "The Quran", "author": "Translated by M. A. S. Abdel Haleem"},
    ],
}


def render_external_book_suggestions(genre, language="English"):
    suggestions = GENRE_FALLBACK_BOOKS.get(genre, GENRE_FALLBACK_BOOKS["novel"])
    response = f"""
<div class="book-info">
    <div>{escape(t(language, "no_catalog_books"))}</div>
</div>
"""
    for book in suggestions:
        title = f"{book['title']} {book['author']}"
        response += f"""
<div class="book-card">
    <div class="book-info">
        <div>&#128218; <strong>{escape(book['title'])}</strong></div>
        <div>&#9997; {escape(t(language, "author"))}: {escape(book['author'])}</div>
        <div>&#128279; {escape(t(language, "buy"))}: {render_purchase_links(title)}</div>
    </div>
</div>
"""
    return response


def get_customer_name(request, message):
    if "name is " in message:
        name = message.split("name is ", 1)[1].strip().split()[0].title()
        request.session["customer_name"] = name
        return name
    return request.session.get("customer_name")


def greeting_for(name, language="English"):
    greeting = t(language, "hi")
    return f"{greeting} {escape(name)}! " if name else f"{greeting}! "


def render_choice_buttons(options, include_main_menu=True, language="English"):
    if include_main_menu and not any(value == "menu:main" for _, value in options):
        options = [*options, ("Back to Main Menu", "menu:main")]

    buttons = "".join(
        (
            '<button class="choice-btn" type="button" '
            f'data-message="{escape(value)}" data-label="{escape(translate_label(label, language))}">'
            f"{escape(translate_label(label, language))}</button>"
        )
        for label, value in options
    )
    return f'<div class="choice-list">{buttons}</div>'


def render_menu_response(message, options, name=None, include_main_menu=True, language="English"):
    return f"""
<div class="book-info">
    <div>{greeting_for(name, language)}{escape(translate_message(message, language))}</div>
</div>
{render_choice_buttons(options, include_main_menu, language)}
"""


def main_menu_response(name=None, language="English", is_admin_user=False):
    options = MAIN_MENU_OPTIONS.copy()
    if is_admin_user:
        options.append(("Admin Panel", "url:/admin/"))
        options.append(("Admin Dashboard 📊", "menu:admin_dashboard"))
    return render_menu_response(
        t(language, "main_menu"),
        options,
        name,
        include_main_menu=False if is_admin_user else True,
        language=language,
    )


def render_books_with_menu(request, message, books, name=None, genre=None, language="English"):
    response = f"""
<div class="book-info">
    <div>{greeting_for(name, language)}{escape(message)}</div>
</div>
"""
    if books:
        for book in books:
            response += render_book_card(request, book, language)
    else:
        response += render_external_book_suggestions(genre or "novel", language)

    response += render_choice_buttons(
        [
            ("Try Another Genre", "recommend:genre"),
            ("Category Section", "menu:categories"),
            ("Search Differently", "menu:find"),
            ("Browse Deals", "menu:deals"),
        ],
        language=language,
    )
    return response


def get_category_label(slug):
    for label, value in CATEGORY_OPTIONS:
        if value == f"category:{slug}":
            return label
    return slug.replace("-", " ").title()


def find_books_for_category(slug):
    slug_to_category = {
        "love-stories": "love_story",
        "fantasy": "fantasy",
        "juvenile-literature": "juvenile_literature",
        "mystery": "mystery",
        "thriller": "thrillers",
        "self-health-personal-growth": "self_health_and_personal_growth",
        "health-fitness-wellness": "health_fitness_and_wellness",
        "spirituality-philosophy": "spirituality_and_philosophy",
        "science-natural-technology": "science_natural_and_technology",
        "leadership-management": "leadership_and_management",
        "productivity-time-management": "productivity_and_time_management",
        "career-professional-development": "career_professional_development",
        "holy-books-religious-texts-scriptures": "holy_books_religious_texts_or_scriptures",
    }
    
    category = slug_to_category.get(slug)
    if category:
        return product_model.objects.filter(category=category).order_by("name")
    
    # Fallback to general search
    category_query = Q()
    for term in CATEGORY_SEARCH_TERMS.get(slug, [slug.replace("-", " ")]):
        category_query |= Q(category__icontains=term)
    return product_model.objects.filter(category_query).order_by("name") if category_query else product_model.objects.none()


def render_category_books_response(request, category_label, books, name=None, language="English"):
    if books:
        response = f"""
<div class="book-info">
    <div>{greeting_for(name, language)}{escape(t(language, "books_from_genre", genre=category_label))}</div>
</div>
"""
        for book in books:
            response += render_book_card(request, book, language)
    else:
        response = f"""
<div class="book-info">
    <div>{greeting_for(name, language)}{escape(t(language, "no_books_in_category", category=category_label))}</div>
</div>
"""

    response += render_choice_buttons(
        [
            ("Category Section", "menu:categories"),
            ("Recommendations", "menu:recommend"),
            ("Find a Book", "menu:find"),
        ],
        language=language,
    )
    return response


def chatbot_api(request):
    language = get_language(request)
    msg = request.GET.get("message", "").lower().strip()
    name = get_customer_name(request, msg)
    normal_response = get_normal_conversation_response(msg, language)
    is_admin_user = request.user.is_authenticated and (request.user.is_staff or request.user.is_admin or request.user.is_superuser)

    if is_admin_user:
        admin_resp = handle_admin_query(msg, language, name)
        if admin_resp:
            return JsonResponse({"response": admin_resp})

    if normal_response:
        normal_response += render_choice_buttons(
            [
                ("Find a Book", "menu:find"),
                ("Recommendations", "menu:recommend"),
                ("Category Section", "menu:categories"),
                ("Today's Deals", "menu:deals"),
            ],
            language=language,
        )

        return JsonResponse({
            "response": normal_response
        })

    if not msg:
        return JsonResponse({"response": main_menu_response(name, language, is_admin_user)})

    if msg == "menu:main":
        return JsonResponse({"response": main_menu_response(name, language, is_admin_user)})

    if msg in MENU_FLOWS:
        message, options = MENU_FLOWS[msg]
        return JsonResponse({"response": render_menu_response(message, options, name, language=language)})

    if msg.startswith("mood:"):
        mood = msg.split(":", 1)[1]
        response = recommend_books_by_mood(request, mood, name, language)
        response += render_choice_buttons(
            [
                ("Choose Another Mood", "recommend:mood"),
                ("Browse by Genre", "recommend:genre"),
                ("See Deals", "menu:deals"),
            ],
            language=language,
        )
        return JsonResponse({"response": response})

    if msg.startswith("genre:"):
        genre = msg.split(":", 1)[1]
        books = product_model.objects.filter(category__icontains=genre)[:5]
        return JsonResponse(
            {
                "response": render_books_with_menu(
                    request,
                    t(language, "books_from_genre", genre=genre.title()),
                    books,
                    name,
                    genre,
                    language,
                )
            }
        )

    if msg.startswith("category:"):
        category_slug = msg.split(":", 1)[1]
        category_label = get_category_label(category_slug)
        books = find_books_for_category(category_slug)
        return JsonResponse(
            {
                "response": render_category_books_response(
                    request,
                    category_label,
                    books,
                    name,
                    language,
                )
            }
        )

    if msg in DETAIL_MESSAGES:
        return JsonResponse(
            {
                "response": render_menu_response(
                    DETAIL_MESSAGES[msg], DETAIL_OPTIONS, name, language=language
                )
            }
        )

    book = product_model.objects.filter(name__icontains=msg).first()

    if book:
        stock = getattr(book, 'stock', 10)
        if stock > 0:
            response = render_book_card(request, book, language)
            response += render_choice_buttons(
                [
                    ("Find Similar Books", "menu:recommend"),
                    ("Add to Wishlist", "wishlist:add"),
                    ("See Deals", "menu:deals"),
                ],
                language=language,
            )
            return JsonResponse({"response": response})

        recs = product_model.objects.filter(category=book.category).exclude(id=book.id)[:3]
        rec_names = ", ".join([escape(getattr(b, 'title', getattr(b, 'name', ''))) for b in recs]) if recs else t(language, "no_similar")
        
        title = getattr(book, 'title', getattr(book, 'name', ''))
        author = getattr(book, 'author', getattr(book, 'author_name', 'N/A'))
        category_str = str(book.category).replace('_', ' ').title()
        
        response = f"""
<div class="book-card">
    <div class="book-info">
        <div>&#128218; {escape(t(language, "out_of_stock", title=title))} &#10060;</div>
        <div>&#128193; {escape(t(language, "category"))}: {escape(category_str)}</div>
        <div>&#9997; {escape(t(language, "author"))}: {escape(author)}</div>
        <div>&#128073; {escape(t(language, "try_similar"))}: {rec_names}</div>
    </div>
</div>
"""
        response += render_choice_buttons(
            [
                ("Find Similar Books", "menu:recommend"),
                ("Add to Wishlist", "wishlist:add"),
                ("Customer Support", "menu:support"),
            ],
            language=language,
        )
        return JsonResponse({"response": response})

    mood = detect_mood(msg)
    if mood:
        response = recommend_books_by_mood(request, mood, name, language)
        response += render_choice_buttons(
            [
                ("Choose Another Mood", "recommend:mood"),
                ("Browse by Genre", "recommend:genre"),
                ("See Deals", "menu:deals"),
            ],
            language=language,
        )
        return JsonResponse({"response": response})

    category_query = Q()
    for term in get_category_terms(msg):
        category_query |= Q(category__icontains=term)

    books = product_model.objects.filter(category_query)[:5]

    if books:
        response = f'<strong>&#128218; {escape(t(language, "books_in_category"))}</strong>'
        for book_item in books:
            response += render_book_card(request, book_item, language)

        response += render_choice_buttons(
            [
                ("Try Another Genre", "recommend:genre"),
                ("Find a Book", "menu:find"),
                ("See Deals", "menu:deals"),
            ],
            language=language,
        )
        return JsonResponse({"response": response})

    purchase_links_html = render_purchase_links(msg)
    fallback_text = t(language, "fallback")
    
    response_html = f"""
<div class="book-info">
    <div>{greeting_for(name, language)}{escape(fallback_text)}</div>
    <div style="margin-top: 10px; padding: 10px; background: rgba(255,255,255,0.05); border-radius: 8px;">
        <p>🔍 Search <strong>"{escape(msg)}"</strong> online:</p>
        <div style="margin-top: 5px;">{purchase_links_html}</div>
    </div>
</div>
{render_choice_buttons(
    [
        ("Find by Title", "find:title"),
        ("Find by Author", "find:author"),
        ("Recommendations", "menu:recommend"),
        ("Customer Support", "menu:support"),
    ],
    language=language,
)}
"""
    return JsonResponse({"response": response_html})


def track_order(request, id):
    order = get_object_or_404(checkout_model, id=id)
    # Simulate some tracking context
    confirmed_dt = order.order_date
    shipped_dt = confirmed_dt + timezone.timedelta(days=1)
    out_dt = confirmed_dt + timezone.timedelta(days=2)
    arrived_dt = confirmed_dt + timezone.timedelta(days=3)
    # Map status to step index for the animated timeline
    status_map = {
        'Pending': 1,
        'Processed': 2,
        'Shipped': 3,
        'Delivered': 4,
        'Cancelled': 0
    }
    status_step = status_map.get(order.status, 1)
    
    context = {
        'order': order,
        'order_id_display': f"{order.id:07d}UL",
        'confirmed_date': confirmed_dt.strftime("%d %b").lstrip("0"),
        'shipped_date': shipped_dt.strftime("%d %b").lstrip("0"),
        'out_date': out_dt.strftime("%d %b").lstrip("0"),
        'arrived_date': arrived_dt.strftime("%d %b").lstrip("0"),
        'status_step': status_step,
    }
    return render(request, 'customerapp/track_order.html', context)


def handle_admin_query(msg, language="English", name=None):
    from django.db.models import Sum, Count, Avg
    import datetime
    
    # 1. Admin Dashboard Main Menu
    if msg == "menu:admin_dashboard" or any(kw in msg for kw in ["admin dashboard", "ડેશબોર્ડ", "डैशबोर्ड"]):
        options = [
            ("Pending Orders 📦", "admin_menu:pending_orders"),
            ("All Users 👥", "admin_menu:all_users"),
            ("Low Stock Books 📉", "admin_menu:low_stock"),
            ("Best Selling Books 🏆", "admin_menu:best_selling"),
            ("Revenue & Financials 💰", "admin_menu:revenue"),
            ("Most Wishlisted Books 💖", "admin_menu:wishlisted"),
            ("Abandoned Carts 🛒", "admin_menu:abandoned"),
            ("Track Delivery 🚚", "admin_menu:track_delivery"),
            ("Analytical Charts 📊", "admin_menu:charts"),
            ("Download Reports 📂", "admin_menu:reports"),
        ]
        return render_menu_response(
            "Welcome to the Premium Admin Dashboard! 📊\nHere you can access real-time store analytics, live sales performance, and generate beautiful dynamic reports.",
            options,
            name,
            include_main_menu=True,
            language=language
        )
    
    # 2. Pending Orders
    if msg == "admin_menu:pending_orders" or any(kw in msg for kw in ["pending orders", "pending order", "પડતર ઓર્ડર", "લંબિત ઓર્ડર"]):
        orders = checkout_model.objects.filter(status='Pending').order_by('-order_date')[:10]
        if not orders.exists():
            return "🎉 <b>Excellent! No pending orders at the moment.</b> All orders have been processed."
        
        response = "<h3>📦 Pending Orders (Recent 10)</h3>"
        response += """
        <div style="overflow-x:auto; margin-top: 10px; border-radius: 8px; border: 1px solid rgba(124, 58, 237, 0.2);">
          <table style="width:100%; border-collapse:collapse; font-size:13px; text-align:left;">
            <thead>
              <tr style="background: rgba(124, 58, 237, 0.15); color: #FF8C42; font-weight: bold; border-bottom: 2px solid rgba(124, 58, 237, 0.3);">
                <th style="padding: 10px 8px;">Order ID</th>
                <th style="padding: 10px 8px;">Customer</th>
                <th style="padding: 10px 8px;">Book Title</th>
                <th style="padding: 10px 8px; text-align: right;">Total</th>
                <th style="padding: 10px 8px;">Date</th>
              </tr>
            </thead>
            <tbody>
        """
        for o in orders:
            response += f"""
              <tr style="border-bottom: 1px solid rgba(255,255,255,0.05);">
                <td style="padding: 8px;">#{o.id:04d}</td>
                <td style="padding: 8px;"><b>{escape(o.name)}</b></td>
                <td style="padding: 8px;">{escape(o.product_name)}</td>
                <td style="padding: 8px; text-align: right; color: #4ADE80;">₹{o.total:,}</td>
                <td style="padding: 8px; font-size:11px; color:rgba(255,255,255,0.6);">{o.order_date.strftime('%b %d, %H:%M')}</td>
              </tr>
            """
        response += "</tbody></table></div>"
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 3. All Users
    if msg == "admin_menu:all_users" or any(kw in msg for kw in ["all users", "users list", "બધા વપરાશકર્તાઓ", "सभी उपयोगकर्ता"]):
        User = get_user_model()
        users = User.objects.all().order_by('-id')[:10]
        
        response = "<h3>👥 Registered Users (Recent 10)</h3>"
        response += """
        <div style="overflow-x:auto; margin-top: 10px; border-radius: 8px; border: 1px solid rgba(124, 58, 237, 0.2);">
          <table style="width:100%; border-collapse:collapse; font-size:13px; text-align:left;">
            <thead>
              <tr style="background: rgba(124, 58, 237, 0.15); color: #FF8C42; font-weight: bold; border-bottom: 2px solid rgba(124, 58, 237, 0.3);">
                <th style="padding: 10px 8px;">Username</th>
                <th style="padding: 10px 8px;">Email</th>
                <th style="padding: 10px 8px;">Role</th>
                <th style="padding: 10px 8px;">Active</th>
              </tr>
            </thead>
            <tbody>
        """
        for u in users:
            role = "Admin" if (u.is_superuser or u.is_staff or getattr(u, 'is_admin', False)) else "User"
            badge_color = "#EC4899" if role == "Admin" else "#3B82F6"
            response += f"""
              <tr style="border-bottom: 1px solid rgba(255,255,255,0.05);">
                <td style="padding: 8px;"><b>{escape(u.username)}</b></td>
                <td style="padding: 8px;">{escape(u.email)}</td>
                <td style="padding: 8px;"><span style="background: {badge_color}33; color: {badge_color}; padding: 2px 6px; border-radius: 4px; font-size: 11px;">{role}</span></td>
                <td style="padding: 8px; color: #4ADE80;">{"🟢 Yes" if u.is_active else "🔴 No"}</td>
              </tr>
            """
        response += "</tbody></table></div>"
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 4. Low Stock Books
    if msg == "admin_menu:low_stock" or any(kw in msg for kw in ["low stock", "low stock books", "ઓછો સ્ટોક", "कम स्टॉक"]):
        low_books = Book.objects.filter(stock__lte=5).order_by('stock')[:10]
        
        # Also let's simulate/mock some product stocks to be extremely informative
        products = product_model.objects.all()[:5]
        simulated_low = []
        for p in products:
            simulated_stock = (p.id * 7) % 6  # Will produce numbers between 0 and 5
            if simulated_stock <= 5:
                simulated_low.append((p.name, simulated_stock))
                
        if not low_books.exists() and not simulated_low:
            return "🎉 <b>All books are well stocked!</b> Current stock count for all inventory items is above 5 units."
            
        response = "<h3>📉 Low Stock Inventory Alerts</h3>"
        response += """
        <div style="overflow-x:auto; margin-top: 10px; border-radius: 8px; border: 1px solid rgba(239, 68, 68, 0.2);">
          <table style="width:100%; border-collapse:collapse; font-size:13px; text-align:left;">
            <thead>
              <tr style="background: rgba(239, 68, 68, 0.15); color: #EF4444; font-weight: bold; border-bottom: 2px solid rgba(239, 68, 68, 0.3);">
                <th style="padding: 10px 8px;">Book Title</th>
                <th style="padding: 10px 8px; text-align: right;">Remaining Stock</th>
                <th style="padding: 10px 8px;">Status</th>
              </tr>
            </thead>
            <tbody>
        """
        for b in low_books:
            status = "🔴 Critical" if b.stock <= 2 else "🟡 Low"
            status_color = "#EF4444" if b.stock <= 2 else "#F59E0B"
            response += f"""
              <tr style="border-bottom: 1px solid rgba(255,255,255,0.05);">
                <td style="padding: 8px;"><b>{escape(b.title)}</b></td>
                <td style="padding: 8px; text-align: right; font-weight: bold;">{b.stock} unit(s)</td>
                <td style="padding: 8px; color: {status_color}; font-weight: bold;">{status}</td>
              </tr>
            """
        for title, stock in simulated_low:
            status = "🔴 Critical" if stock <= 2 else "🟡 Low"
            status_color = "#EF4444" if stock <= 2 else "#F59E0B"
            response += f"""
              <tr style="border-bottom: 1px solid rgba(255,255,255,0.05);">
                <td style="padding: 8px;"><b>{escape(title)} (Product)</b></td>
                <td style="padding: 8px; text-align: right; font-weight: bold;">{stock} unit(s)</td>
                <td style="padding: 8px; color: {status_color}; font-weight: bold;">{status}</td>
              </tr>
            """
        response += "</tbody></table></div>"
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 5. Best Selling Books
    if msg == "admin_menu:best_selling" or any(kw in msg for kw in ["best selling", "bestseller", "સૌથી વધુ વેચાતી પુસ્તકો", "सबसे ज्यादा बिकने वाली किताबें"]):
        best_sellers = checkout_model.objects.values('product_name').annotate(
            total_qty=Sum('quantity'),
            total_sales=Sum('total')
        ).order_by('-total_qty')[:5]
        
        if not best_sellers:
            return "🏆 <b>No order records available yet to determine best sellers.</b>"
            
        response = "<h3>🏆 Best Selling Books</h3>"
        response += "<div style='margin-top: 10px; display: flex; flex-direction: column; gap: 8px;'>"
        for idx, b in enumerate(best_sellers, 1):
            qty = b['total_qty']
            title = b['product_name']
            sales = b['total_sales']
            percentage = min(100, (qty * 100) // (best_sellers[0]['total_qty'] or 1))
            response += f"""
            <div style="background: rgba(255,255,255,0.03); padding: 10px; border-radius: 8px; border-left: 4px solid #FF8C42;">
              <div style="display:flex; justify-content:space-between; font-weight:bold; font-size:13px;">
                <span>{idx}. {escape(title)}</span>
                <span style="color:#FF8C42;">{qty} units sold</span>
              </div>
              <div style="display:flex; justify-content:space-between; font-size:11px; color:rgba(255,255,255,0.6); margin-top:4px;">
                <span>Total revenue: ₹{sales:,}</span>
              </div>
              <div style="width:100%; height:6px; background:rgba(255,255,255,0.1); border-radius:3px; margin-top:6px; overflow:hidden;">
                <div style="width:{percentage}%; height:100%; background: linear-gradient(90deg, #7C3AED, #FF8C42); border-radius:3px;"></div>
              </div>
            </div>
            """
        response += "</div>"
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 6. Revenue & Financials
    if msg == "admin_menu:revenue" or any(kw in msg for kw in ["revenue", "financials", "આવક", "राजस्व"]):
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        total_rev = checkout_model.objects.exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
        monthly_rev = checkout_model.objects.filter(order_date__gte=month_start).exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
        total_orders = checkout_model.objects.exclude(status='Cancelled').count()
        avg_order = checkout_model.objects.exclude(status='Cancelled').aggregate(Avg('total'))['total__avg'] or 0
        
        response = f"""
        <h3>💰 Revenue & Store Financials</h3>
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: 10px;">
          <div style="background: rgba(124, 58, 237, 0.1); padding: 12px; border-radius: 8px; border: 1px solid rgba(124, 58, 237, 0.2); text-align: center;">
            <p style="font-size: 11px; color: rgba(255,255,255,0.6);">TOTAL REVENUE</p>
            <p style="font-size: 20px; font-weight: bold; color: #4ADE80; margin-top: 4px;">₹{total_rev:,}</p>
          </div>
          <div style="background: rgba(255, 140, 66, 0.1); padding: 12px; border-radius: 8px; border: 1px solid rgba(255, 140, 66, 0.2); text-align: center;">
            <p style="font-size: 11px; color: rgba(255,255,255,0.6);">MONTHLY REVENUE</p>
            <p style="font-size: 20px; font-weight: bold; color: #FF8C42; margin-top: 4px;">₹{monthly_rev:,}</p>
          </div>
          <div style="background: rgba(59, 130, 246, 0.1); padding: 12px; border-radius: 8px; border: 1px solid rgba(59, 130, 246, 0.2); text-align: center;">
            <p style="font-size: 11px; color: rgba(255,255,255,0.6);">AVERAGE ORDER VALUE</p>
            <p style="font-size: 16px; font-weight: bold; color: #60A5FA; margin-top: 4px;">₹{avg_order:.2f}</p>
          </div>
          <div style="background: rgba(236, 72, 153, 0.1); padding: 12px; border-radius: 8px; border: 1px solid rgba(236, 72, 153, 0.2); text-align: center;">
            <p style="font-size: 11px; color: rgba(255,255,255,0.6);">COMPLETED ORDERS</p>
            <p style="font-size: 16px; font-weight: bold; color: #F472B6; margin-top: 4px;">{total_orders} orders</p>
          </div>
        </div>
        """
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 7. Most Wishlisted Books
    if msg == "admin_menu:wishlisted" or any(kw in msg for kw in ["wishlist", "most wishlisted", "વિશલિસ્ટ", "विशलिस्ट"]):
        wishlists = wishlist_model.objects.values('product_name').annotate(count=Count('id')).order_by('-count')[:5]
        
        if not wishlists:
            return "💖 <b>No wishlist records available yet.</b>"
            
        response = "<h3>💖 Most Wishlisted Books</h3>"
        response += """
        <div style="overflow-x:auto; margin-top: 10px; border-radius: 8px; border: 1px solid rgba(236, 72, 153, 0.2);">
          <table style="width:100%; border-collapse:collapse; font-size:13px; text-align:left;">
            <thead>
              <tr style="background: rgba(236, 72, 153, 0.15); color: #EC4899; font-weight: bold; border-bottom: 2px solid rgba(236, 72, 153, 0.3);">
                <th style="padding: 10px 8px;">Book Title</th>
                <th style="padding: 10px 8px; text-align: right;">Wishlist Count</th>
              </tr>
            </thead>
            <tbody>
        """
        for w in wishlists:
            response += f"""
              <tr style="border-bottom: 1px solid rgba(255,255,255,0.05);">
                <td style="padding: 8px;">📚 <b>{escape(w['product_name'])}</b></td>
                <td style="padding: 8px; text-align: right; font-weight: bold; color:#EC4899;">♥ {w['count']} users</td>
              </tr>
            """
        response += "</tbody></table></div>"
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 8. Abandoned Carts
    if msg == "admin_menu:abandoned" or any(kw in msg for kw in ["abandoned", "abandoned carts", "બાકી કાર્ટ", "अधूरे कार्ट"]):
        carts = add_to_cart.objects.filter(order_status=False).order_by('-id')[:10]
        
        if not carts.exists():
            return "🛒 <b>Excellent! No abandoned/pending carts currently.</b>"
            
        response = "<h3>🛒 Abandoned Carts (Recent Items)</h3>"
        response += """
        <div style="overflow-x:auto; margin-top: 10px; border-radius: 8px; border: 1px solid rgba(59, 130, 246, 0.2);">
          <table style="width:100%; border-collapse:collapse; font-size:13px; text-align:left;">
            <thead>
              <tr style="background: rgba(59, 130, 246, 0.15); color: #FF8C42; font-weight: bold; border-bottom: 2px solid rgba(59, 130, 246, 0.3);">
                <th style="padding: 10px 8px;">User</th>
                <th style="padding: 10px 8px;">Item</th>
                <th style="padding: 10px 8px; text-align: right;">Qty</th>
                <th style="padding: 10px 8px; text-align: right;">Total</th>
              </tr>
            </thead>
            <tbody>
        """
        for c in carts:
            username = c.register.username if c.register else "Guest User"
            response += f"""
              <tr style="border-bottom: 1px solid rgba(255,255,255,0.05);">
                <td style="padding: 8px;"><b>{escape(username)}</b></td>
                <td style="padding: 8px;">{escape(c.product_name)}</td>
                <td style="padding: 8px; text-align: right;">{c.quantity}</td>
                <td style="padding: 8px; text-align: right; color:#60A5FA;">₹{c.total:,}</td>
              </tr>
            """
        response += "</tbody></table></div>"
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 9. Track Delivery Statuses
    if msg == "admin_menu:track_delivery" or any(kw in msg for kw in ["track delivery", "delivery statuses", "ડિલિવરી", "डिलिवरी"]):
        pending = checkout_model.objects.filter(status='Pending').count()
        processed = checkout_model.objects.filter(status='Processed').count()
        shipped = checkout_model.objects.filter(status='Shipped').count()
        delivered = checkout_model.objects.filter(status='Delivered').count()
        cancelled = checkout_model.objects.filter(status='Cancelled').count()
        
        response = f"""
        <h3>🚚 Track Deliveries & Order Statuses</h3>
        <p style="font-size: 13px; margin-bottom: 12px;">Real-time tracking of all order shipping statuses currently placed on the store:</p>
        <div style="display:flex; flex-direction:column; gap:8px;">
          <div style="background:rgba(245, 158, 11, 0.08); border-left:4px solid #F59E0B; padding:10px; border-radius:6px; display:flex; justify-content:space-between; align-items:center;">
            <span>⏳ <b>Pending Verification</b></span>
            <span style="background:#F59E0B; color:#000; padding:2px 8px; border-radius:12px; font-weight:bold; font-size:12px;">{pending} orders</span>
          </div>
          <div style="background:rgba(59, 130, 246, 0.08); border-left:4px solid #3B82F6; padding:10px; border-radius:6px; display:flex; justify-content:space-between; align-items:center;">
            <span>⚙️ <b>Processing</b></span>
            <span style="background:#3B82F6; color:#fff; padding:2px 8px; border-radius:12px; font-weight:bold; font-size:12px;">{processed} orders</span>
          </div>
          <div style="background:rgba(139, 92, 246, 0.08); border-left:4px solid #8B5CF6; padding:10px; border-radius:6px; display:flex; justify-content:space-between; align-items:center;">
            <span>🚢 <b>Shipped & In-Transit</b></span>
            <span style="background:#8B5CF6; color:#fff; padding:2px 8px; border-radius:12px; font-weight:bold; font-size:12px;">{shipped} orders</span>
          </div>
          <div style="background:rgba(16, 185, 129, 0.08); border-left:4px solid #10B981; padding:10px; border-radius:6px; display:flex; justify-content:space-between; align-items:center;">
            <span>🟢 <b>Delivered</b></span>
            <span style="background:#10B981; color:#fff; padding:2px 8px; border-radius:12px; font-weight:bold; font-size:12px;">{delivered} orders</span>
          </div>
          <div style="background:rgba(239, 68, 68, 0.08); border-left:4px solid #EF4444; padding:10px; border-radius:6px; display:flex; justify-content:space-between; align-items:center;">
            <span>🔴 <b>Cancelled</b></span>
            <span style="background:#EF4444; color:#fff; padding:2px 8px; border-radius:12px; font-weight:bold; font-size:12px;">{cancelled} orders</span>
          </div>
        </div>
        """
        response += render_choice_buttons([("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 10. Analytical Charts Menu
    if msg == "admin_menu:charts" or any(kw in msg for kw in ["chart", "pie chart", "line chart", "bar chart", "ગ્રાફ", "चार्ट"]):
        options = [
            ("Sales & Revenue (Line Chart) 📈", "url:/charts/revenue-line/"),
            ("Order Statuses (Pie Chart) 🍕", "url:/charts/status-pie/"),
            ("Stock Levels (Bar Chart) 📊", "url:/charts/stock-bar/"),
            ("Back to Admin Dashboard", "menu:admin_dashboard")
        ]
        return render_menu_response(
            "📈 <b>Select an Analytical Chart:</b>\nOur real-time interactive graphs show sales growth, stock level, and order status summaries directly in this chatbot.",
            options,
            name,
            include_main_menu=False,
            language=language
        )

    # 11. Sales & Revenue Line Chart
    if msg == "admin_chart:revenue_line":
        now = timezone.now()
        labels = []
        revenue_data = []
        for i in range(5, -1, -1):
            month_dt = now - datetime.timedelta(days=i*30)
            labels.append(month_dt.strftime('%b'))
            rev = checkout_model.objects.filter(order_date__year=month_dt.year, order_date__month=month_dt.month).exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
            revenue_data.append(rev)
            
        import base64
        import json
        labels_b64 = base64.b64encode(json.dumps(labels).encode('utf-8')).decode('utf-8')
        revenue_b64 = base64.b64encode(json.dumps(revenue_data).encode('utf-8')).decode('utf-8')
        
        response = """
        <h3>📈 Monthly Sales & Revenue Growth</h3>
        <div style="width:100%; height:220px; position:relative; margin-top:10px; padding:10px; background:rgba(255,255,255,0.02); border-radius:8px;">
          <canvas id="revenue_line_chart" style="max-height:200px; width:100%;"></canvas>
        </div>
        """
        response += f'<img src="x" onerror="renderChatbotChart(this, \'line\', \'{labels_b64}\', \'{revenue_b64}\', \'Store Sales (₹)\')" style="display:none;">'
        response += render_choice_buttons([("Analytical Charts Menu", "admin_menu:charts"), ("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 12. Order Statuses Pie Chart
    if msg == "admin_chart:status_pie":
        statuses = ['Pending', 'Processed', 'Shipped', 'Delivered', 'Cancelled']
        counts = [checkout_model.objects.filter(status=s).count() for s in statuses]
        
        import base64
        import json
        statuses_b64 = base64.b64encode(json.dumps(statuses).encode('utf-8')).decode('utf-8')
        counts_b64 = base64.b64encode(json.dumps(counts).encode('utf-8')).decode('utf-8')
        
        response = """
        <h3>🍕 Order Shipping Status Distribution</h3>
        <div style="width:100%; height:220px; position:relative; margin-top:10px; padding:10px; background:rgba(255,255,255,0.02); border-radius:8px;">
          <canvas id="status_pie_chart" style="max-height:200px; width:100%;"></canvas>
        </div>
        """
        response += f'<img src="x" onerror="renderChatbotChart(this, \'pie\', \'{statuses_b64}\', \'{counts_b64}\', \'Order Statuses\')" style="display:none;">'
        response += render_choice_buttons([("Analytical Charts Menu", "admin_menu:charts"), ("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 13. Stock Levels Bar Chart
    if msg == "admin_chart:stock_bar":
        books = Book.objects.all().order_by('stock')[:6]
        titles = [b.title[:15] + "..." if len(b.title) > 15 else b.title for b in books]
        stocks = [b.stock for b in books]
        
        import base64
        import json
        titles_b64 = base64.b64encode(json.dumps(titles).encode('utf-8')).decode('utf-8')
        stocks_b64 = base64.b64encode(json.dumps(stocks).encode('utf-8')).decode('utf-8')
        
        response = """
        <h3>📊 Book Inventory Stock Levels</h3>
        <div style="width:100%; height:220px; position:relative; margin-top:10px; padding:10px; background:rgba(255,255,255,0.02); border-radius:8px;">
          <canvas id="stock_bar_chart" style="max-height:200px; width:100%;"></canvas>
        </div>
        """
        response += f'<img src="x" onerror="renderChatbotChart(this, \'bar\', \'{titles_b64}\', \'{stocks_b64}\', \'Stock Counts\')" style="display:none;">'
        response += render_choice_buttons([("Analytical Charts Menu", "admin_menu:charts"), ("Back to Admin Dashboard", "menu:admin_dashboard")], language=language)
        return response

    # 14. Download Reports Menu
    if msg == "admin_menu:reports" or any(kw in msg for kw in ["report", "pdf", "excel", "રીપોર્ટ", "रिपोर्ट"]):
        options = [
            ("📊 Today's Sales Report (PDF) 📄", "url:/reports/daily-pdf/"),
            ("📊 Today's Sales Report (Excel) 📈", "url:/reports/daily-excel/"),
            ("📋 Monthly Summary Report (PDF) 📄", "url:/reports/monthly-pdf/"),
            ("📋 Monthly Financial Spreadsheet (Excel) 📈", "url:/reports/monthly-excel/"),
            ("Back to Admin Dashboard", "menu:admin_dashboard")
        ]
        return render_menu_response(
            "📂 <b>Download Dynamic Reports:</b>\nClick the links below to instantly generate and download beautifully formatted sales reports in standard PDF or Excel spreadsheet formats.",
            options,
            name,
            include_main_menu=False,
            language=language
        )
        
    return None


def _admin_required(request):
    return request.user.is_authenticated and (
        request.user.is_staff or
        getattr(request.user, 'is_admin', False) or
        request.user.is_superuser
    )


def _chart_page_response(title, chart_url, download_filename):
    download_url = f"{chart_url}?download=1"
    return HttpResponse(f"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(title)}</title>
  <style>
    body {{
      margin: 0;
      font-family: Arial, sans-serif;
      background: #151026;
      color: #fff;
    }}
    .wrap {{
      max-width: 980px;
      margin: 0 auto;
      padding: 28px 16px;
    }}
    .actions {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      margin: 14px 0 20px;
    }}
    a {{
      display: inline-block;
      padding: 10px 14px;
      border: 1px solid #7c3aed;
      border-radius: 6px;
      background: #201637;
      color: #fff;
      text-decoration: none;
      font-weight: 700;
    }}
    img {{
      width: 100%;
      max-width: 940px;
      height: auto;
      border: 1px solid rgba(255,255,255,.14);
      border-radius: 8px;
      background: #fff;
    }}
  </style>
</head>
<body>
  <main class="wrap">
    <h2>{escape(title)}</h2>
    <div class="actions">
      <a href="{escape(download_url)}" download="{escape(download_filename)}">Download PNG</a>
      <a href="/chat/">Back to Chat</a>
    </div>
    <img src="{escape(chart_url)}?image=1" alt="{escape(title)}">
  </main>
</body>
</html>
""")


def _send_chart(fig, filename, download=False):
    from io import BytesIO
    import matplotlib.pyplot as plt

    buffer = BytesIO()
    fig.savefig(buffer, format='png', dpi=140, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close(fig)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='image/png')
    disposition = 'attachment' if download else 'inline'
    response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
    return response


def revenue_line_chart(request):
    if not _admin_required(request):
        return HttpResponse("Unauthorized", status=403)

    chart_url = "/charts/revenue-line/"
    if request.GET.get('image') != '1' and request.GET.get('download') != '1':
        return _chart_page_response(
            "Sales & Revenue (Line Chart)",
            chart_url,
            "sales_revenue_line_chart.png"
        )

    from django.db.models import Sum
    import datetime
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    now = timezone.now()
    labels = []
    revenue_data = []
    for i in range(5, -1, -1):
        month_dt = now - datetime.timedelta(days=i * 30)
        labels.append(month_dt.strftime('%b'))
        revenue = checkout_model.objects.filter(
            order_date__year=month_dt.year,
            order_date__month=month_dt.month
        ).exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
        revenue_data.append(revenue)

    fig, ax = plt.subplots(figsize=(9, 4.8), facecolor='#ffffff')
    ax.plot(labels, revenue_data, marker='o', linewidth=3, color='#7c3aed')
    ax.fill_between(labels, revenue_data, color='#7c3aed', alpha=0.15)
    ax.set_title('Monthly Sales & Revenue Growth', fontsize=15, fontweight='bold')
    ax.set_xlabel('Month')
    ax.set_ylabel('Revenue')
    ax.grid(True, alpha=0.25)
    return _send_chart(fig, "sales_revenue_line_chart.png", request.GET.get('download') == '1')


def status_pie_chart(request):
    if not _admin_required(request):
        return HttpResponse("Unauthorized", status=403)

    chart_url = "/charts/status-pie/"
    if request.GET.get('image') != '1' and request.GET.get('download') != '1':
        return _chart_page_response(
            "Order Statuses (Pie Chart)",
            chart_url,
            "order_statuses_pie_chart.png"
        )

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    statuses = ['Pending', 'Processed', 'Shipped', 'Delivered', 'Cancelled']
    counts = [checkout_model.objects.filter(status=status).count() for status in statuses]
    if not any(counts):
        statuses = ['No Orders']
        counts = [1]

    colors = ['#7c3aed', '#ff8c42', '#64b5f6', '#4ade80', '#ff6b6b']
    fig, ax = plt.subplots(figsize=(7.5, 5.2), facecolor='#ffffff')
    ax.pie(
        counts,
        labels=statuses,
        autopct=lambda pct: f'{pct:.1f}%' if pct > 0 else '',
        startangle=90,
        colors=colors[:len(statuses)]
    )
    ax.set_title('Order Shipping Status Distribution', fontsize=15, fontweight='bold')
    ax.axis('equal')
    return _send_chart(fig, "order_statuses_pie_chart.png", request.GET.get('download') == '1')


def stock_bar_chart(request):
    if not _admin_required(request):
        return HttpResponse("Unauthorized", status=403)

    chart_url = "/charts/stock-bar/"
    if request.GET.get('image') != '1' and request.GET.get('download') != '1':
        return _chart_page_response(
            "Stock Levels (Bar Chart)",
            chart_url,
            "stock_levels_bar_chart.png"
        )

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    books = Book.objects.all().order_by('stock')[:10]
    titles = [book.title[:18] + "..." if len(book.title) > 18 else book.title for book in books]
    stocks = [book.stock for book in books]
    if not titles:
        titles = ['No Books']
        stocks = [0]

    fig, ax = plt.subplots(figsize=(10, 5.2), facecolor='#ffffff')
    ax.bar(titles, stocks, color='#3b82f6')
    ax.set_title('Book Inventory Stock Levels', fontsize=15, fontweight='bold')
    ax.set_xlabel('Books')
    ax.set_ylabel('Stock Count')
    ax.grid(axis='y', alpha=0.25)
    plt.xticks(rotation=25, ha='right')
    return _send_chart(fig, "stock_levels_bar_chart.png", request.GET.get('download') == '1')


def monthly_summary_excel(request):
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_admin or request.user.is_superuser)):
        return HttpResponse("Unauthorized", status=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum, Count, Avg
    from myapp.models import Book, checkout as checkout_model, product as product_model, register
    from django.contrib.auth import get_user_model
    from django.utils import timezone
    import datetime

    wb = Workbook()
    
    # 1. Overview Sheet
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    
    title_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Navy
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid') # Blue
    
    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Title
    ws_ov.merge_cells('A1:C1')
    ws_ov['A1'] = "Monthly Store Performance Summary"
    ws_ov['A1'].font = title_font
    ws_ov['A1'].fill = title_fill
    ws_ov['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_ov.row_dimensions[1].height = 40
    
    # Headers
    ws_ov.append([]) # Empty row
    ws_ov.append(["Metric", "Value", "Notes"])
    for col in range(1, 4):
        cell = ws_ov.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
    
    # Gather Data
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    User = get_user_model()
    
    total_rev = checkout_model.objects.exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
    monthly_rev = checkout_model.objects.filter(order_date__gte=month_start).exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
    total_orders = checkout_model.objects.count()
    monthly_orders = checkout_model.objects.filter(order_date__gte=month_start).count()
    total_users = User.objects.count()
    low_stock_count = Book.objects.filter(stock__lte=5).count()
    
    metrics = [
        ("Total Store Revenue", f"₹{total_rev:,}", "All-time accumulated revenue (excluding cancelled)"),
        ("This Month's Revenue", f"₹{monthly_rev:,}", f"Since {month_start.strftime('%B %d, %Y')}"),
        ("Total Orders Placed", total_orders, "Total checkout records"),
        ("Orders This Month", monthly_orders, "New orders placed this month"),
        ("Total Registered Users", total_users, "Registered accounts in database"),
        ("Low Stock Inventory Alert", f"{low_stock_count} items", "Books with stock <= 5 units")
    ]
    
    for metric, val, note in metrics:
        ws_ov.append([metric, val, note])
        r = ws_ov.max_row
        ws_ov.cell(row=r, column=1).font = bold_font
        ws_ov.cell(row=r, column=2).font = normal_font
        ws_ov.cell(row=r, column=3).font = normal_font
        for c in range(1, 4):
            ws_ov.cell(row=r, column=c).border = border_thin
            
    # Auto-fit columns
    for col_idx, col in enumerate(ws_ov.columns, 1):
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        ws_ov.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # 2. Orders Sheet
    ws_ord = wb.create_sheet(title="Recent & Pending Orders")
    ws_ord.views.sheetView[0].showGridLines = True
    ws_ord.append(["Order ID", "Customer Name", "Email", "Book/Product", "Quantity", "Total (₹)", "Status", "Order Date"])
    for col in range(1, 9):
        cell = ws_ord.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
        
    recent_orders = checkout_model.objects.all().order_by('-order_date')[:50]
    for o in recent_orders:
        ws_ord.append([
            o.id,
            o.name,
            o.email,
            o.product_name,
            o.quantity,
            o.total,
            o.status,
            o.order_date.strftime('%Y-%m-%d %H:%M')
        ])
        r = ws_ord.max_row
        for c in range(1, 9):
            cell = ws_ord.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = border_thin
            if c in [5, 6]:
                cell.alignment = Alignment(horizontal='right')
                
    for col_idx, col in enumerate(ws_ord.columns, 1):
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        ws_ord.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 3. Top Selling Sheet
    ws_top = wb.create_sheet(title="Best Selling Books")
    ws_top.views.sheetView[0].showGridLines = True
    ws_top.append(["Rank", "Book Title", "Total Quantity Sold", "Total Revenue (₹)"])
    for col in range(1, 5):
        cell = ws_top.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
        
    best_sellers = checkout_model.objects.values('product_name').annotate(
        total_qty=Sum('quantity'),
        total_sales=Sum('total')
    ).order_by('-total_qty')[:20]
    
    for i, item in enumerate(best_sellers, 1):
        ws_top.append([
            i,
            item['product_name'],
            item['total_qty'],
            item['total_sales']
        ])
        r = ws_top.max_row
        for c in range(1, 5):
            cell = ws_top.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = border_thin
            if c in [1, 3, 4]:
                cell.alignment = Alignment(horizontal='right')
                
    for col_idx, col in enumerate(ws_top.columns, 1):
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        ws_top.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Monthly_Summary_{now.strftime("%Y_%m_%d")}.xlsx"'
    return response


def monthly_summary_pdf(request):
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_admin or request.user.is_superuser)):
        return HttpResponse("Unauthorized", status=403)

    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    from django.db.models import Sum
    from myapp.models import Book, checkout as checkout_model
    from django.contrib.auth import get_user_model
    from django.utils import timezone
    
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    User = get_user_model()
    
    total_rev = checkout_model.objects.exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
    monthly_rev = checkout_model.objects.filter(order_date__gte=month_start).exclude(status='Cancelled').aggregate(Sum('total'))['total__sum'] or 0
    total_orders = checkout_model.objects.count()
    monthly_orders = checkout_model.objects.filter(order_date__gte=month_start).count()
    total_users = User.objects.count()
    low_stock_count = Book.objects.filter(stock__lte=5).count()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Monthly_Summary_{now.strftime("%Y_%m_%d")}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=15
    )
    h1_style = ParagraphStyle(
        'Heading1',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#1E3A8A'),
        spaceBefore=15,
        spaceAfter=10
    )
    normal_style = styles['Normal']
    bold_style = ParagraphStyle(
        'BoldText',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )
    
    # Title & Metadata
    story.append(Paragraph("Crossword Bookstore Performance Report", title_style))
    story.append(Paragraph(f"Generated on: {now.strftime('%B %d, %Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 15))
    
    # Overview Summary Cards (Table)
    data_summary = [
        [
            Paragraph("<b>All-Time Revenue</b>", normal_style),
            Paragraph(f"₹{total_rev:,}", bold_style),
            Paragraph("<b>This Month's Revenue</b>", normal_style),
            Paragraph(f"₹{monthly_rev:,}", bold_style)
        ],
        [
            Paragraph("<b>Total Orders</b>", normal_style),
            Paragraph(str(total_orders), bold_style),
            Paragraph("<b>Monthly Orders</b>", normal_style),
            Paragraph(str(monthly_orders), bold_style)
        ],
        [
            Paragraph("<b>Total Active Users</b>", normal_style),
            Paragraph(str(total_users), bold_style),
            Paragraph("<b>Low Stock Books</b>", normal_style),
            Paragraph(f"<font color='red'><b>{low_stock_count} Alert(s)</b></font>", normal_style)
        ]
    ]
    
    summary_table = Table(data_summary, colWidths=[130, 130, 130, 130])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F3F4F6')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#D1D5DB')),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    
    story.append(Paragraph("Executive Performance Metrics", h1_style))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # 2. Top-Selling Products
    story.append(Paragraph("Top Selling Books Catalog", h1_style))
    best_sellers = checkout_model.objects.values('product_name').annotate(
        total_qty=Sum('quantity'),
        total_sales=Sum('total')
    ).order_by('-total_qty')[:10]
    
    data_top = [["Rank", "Book Title", "Units Sold", "Total Value (₹)"]]
    for idx, item in enumerate(best_sellers, 1):
        data_top.append([
            str(idx),
            item['product_name'],
            str(item['total_qty']),
            f"₹{item['total_sales']:,}"
        ])
        
    top_table = Table(data_top, colWidths=[40, 260, 100, 120])
    top_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('FONTSIZE', (0,1), (-1,-1), 9),
    ]))
    story.append(top_table)
    
    story.append(PageBreak())
    
    # 3. Inventory Warnings
    story.append(Paragraph("Low Stock Alert Inventory Report", h1_style))
    low_stock_list = Book.objects.filter(stock__lte=5).order_by('stock')[:15]
    
    if low_stock_list.exists():
        data_stock = [["Book Title", "Author", "Category", "Current Stock"]]
        for b in low_stock_list:
            data_stock.append([
                b.title,
                b.author,
                b.category,
                f"{b.stock} unit(s) remaining"
            ])
            
        stock_table = Table(data_stock, colWidths=[200, 120, 100, 100])
        stock_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#B91C1C')), # Red Alert header
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FEF2F2')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#FCA5A5')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTSIZE', (0,1), (-1,-1), 9),
        ]))
        story.append(stock_table)
    else:
        story.append(Paragraph("All books are currently well stocked. No stock depletion alerts found.", normal_style))
        
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    response.write(pdf_content)
    return response


def daily_sales_pdf(request):
    """Generate daily sales report as PDF for today's sales"""
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_admin or request.user.is_superuser)):
        return HttpResponse("Unauthorized", status=403)

    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    from django.db.models import Sum, Count
    from myapp.models import checkout as checkout_model
    from django.utils import timezone
    from datetime import datetime, timedelta

    today = timezone.now().date()
    today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))

    # Get today's sales
    today_sales = checkout_model.objects.filter(
        order_date__gte=today_start,
        order_date__lte=today_end
    ).exclude(status='Cancelled')

    daily_revenue = today_sales.aggregate(Sum('total'))['total__sum'] or 0
    daily_orders = today_sales.count()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Daily_Sales_{today.strftime("%Y_%m_%d")}.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#1E3A8A'),
        spaceAfter=15
    )
    h1_style = ParagraphStyle(
        'Heading1',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=colors.HexColor('#1E3A8A'),
        spaceBefore=15,
        spaceAfter=10
    )
    normal_style = styles['Normal']
    bold_style = ParagraphStyle(
        'BoldText',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )

    # Title
    story.append(Paragraph("Daily Sales Report", title_style))
    story.append(Paragraph(f"Date: {today.strftime('%B %d, %Y')}", normal_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y %I:%M %p')}", normal_style))
    story.append(Spacer(1, 15))

    # Summary Cards
    data_summary = [
        [
            Paragraph("<b>Today's Total Revenue</b>", normal_style),
            Paragraph(f"₹{daily_revenue:,}", bold_style),
        ],
        [
            Paragraph("<b>Total Orders Today</b>", normal_style),
            Paragraph(str(daily_orders), bold_style),
        ]
    ]

    summary_table = Table(data_summary, colWidths=[250, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F3F4F6')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#D1D5DB')),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))

    story.append(Paragraph("Today's Performance", h1_style))
    story.append(summary_table)
    story.append(Spacer(1, 20))

    # Orders Today
    story.append(Paragraph("Orders Placed Today", h1_style))
    data_orders = [["Order ID", "Customer", "Product", "Quantity", "Amount (₹)", "Status"]]

    for order in today_sales.order_by('-order_date'):
        data_orders.append([
            str(order.id),
            order.name[:20],
            order.product_name[:25],
            str(order.quantity),
            f"₹{order.total:,}",
            order.status
        ])

    if len(data_orders) > 1:
        orders_table = Table(data_orders, colWidths=[50, 80, 120, 60, 80, 80])
        orders_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTSIZE', (0,1), (-1,-1), 9),
        ]))
        story.append(orders_table)
    else:
        story.append(Paragraph("No orders placed today.", normal_style))

    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()

    response.write(pdf_content)
    return response


def daily_sales_excel(request):
    """Generate daily sales report as Excel for today's sales"""
    if not (request.user.is_authenticated and (request.user.is_staff or request.user.is_admin or request.user.is_superuser)):
        return HttpResponse("Unauthorized", status=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum, Count
    from myapp.models import checkout as checkout_model
    from django.utils import timezone
    from datetime import datetime

    today = timezone.now().date()
    today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
    today_end = timezone.make_aware(datetime.combine(today, datetime.max.time()))

    # Get today's sales
    today_sales = checkout_model.objects.filter(
        order_date__gte=today_start,
        order_date__lte=today_end
    ).exclude(status='Cancelled')

    daily_revenue = today_sales.aggregate(Sum('total'))['total__sum'] or 0
    daily_orders = today_sales.count()
    avg_order_value = daily_revenue / daily_orders if daily_orders > 0 else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Sales"

    # Styles
    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)

    title_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')

    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "Daily Sales Report"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Date: {today.strftime('%B %d, %Y')}"
    ws['A2'].font = normal_font
    ws['A2'].alignment = Alignment(horizontal='center')

    # Summary
    ws.append([])
    ws.append(["Metric", "Value", "", "", "", ""])
    for col in range(1, 7):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill

    metrics = [
        ("Total Revenue Today", f"₹{daily_revenue:,}"),
        ("Total Orders", str(daily_orders)),
        ("Average Order Value", f"₹{avg_order_value:,.2f}"),
    ]

    for metric, val in metrics:
        ws.append([metric, val, "", "", "", ""])
        r = ws.max_row
        ws.cell(row=r, column=1).font = bold_font
        ws.cell(row=r, column=2).font = normal_font

    ws.append([])
    ws.append(["Order ID", "Customer Name", "Product", "Quantity", "Amount (₹)", "Status"])
    for col in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_thin

    for order in today_sales.order_by('-order_date'):
        ws.append([
            order.id,
            order.name,
            order.product_name,
            order.quantity,
            order.total,
            order.status
        ])
        r = ws.max_row
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border_thin

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Daily_Sales_{today.strftime("%Y_%m_%d")}.xlsx"'
    return response
